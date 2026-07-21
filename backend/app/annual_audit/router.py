"""End-to-end annual-audit project workflow.

The prototype keeps using ObjectInstance as the persistence layer so the new
product flow remains compatible with the ontology explorer and workbench.
"""
from __future__ import annotations

import copy
import hashlib
import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field
from sqlmodel import Session, select

from ..db import DATA_DIR, engine, get_session
from ..models import ObjectInstance, ObjectType


router = APIRouter(prefix="/api/annual-audit", tags=["annual-audit"])

DEMO_ENGAGEMENT_CODE = "ENG-JSDW-2025"
UPLOAD_DIR = DATA_DIR / "annual_audit_uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


class AnnualAuditProjectCreate(BaseModel):
    client_name: str = Field(min_length=2)
    year: int = Field(ge=2000, le=2100)
    period_start: str
    period_end: str
    industry: str = ""
    credit_code: str = ""
    accounting_standard: str
    report_framework: str = "年度财务报表审计"
    materiality_basis: str
    pm: float = Field(gt=0)
    te: float = Field(gt=0)
    trivial_threshold: float = Field(ge=0)
    audit_strategy: str = "综合审计策略"
    first_year: bool = False
    partner: str = ""
    manager: str = ""
    preparer: str = ""
    reviewer: str = ""
    report_date: str = ""
    notes: str = ""
    use_demo_data: bool = False


class AnnualAuditProjectUpdate(BaseModel):
    client_name: str | None = None
    year: int | None = Field(default=None, ge=2000, le=2100)
    period_start: str | None = None
    period_end: str | None = None
    industry: str | None = None
    credit_code: str | None = None
    accounting_standard: str | None = None
    report_framework: str | None = None
    materiality_basis: str | None = None
    pm: float | None = Field(default=None, gt=0)
    te: float | None = Field(default=None, gt=0)
    trivial_threshold: float | None = Field(default=None, ge=0)
    audit_strategy: str | None = None
    first_year: bool | None = None
    partner: str | None = None
    manager: str | None = None
    preparer: str | None = None
    reviewer: str | None = None
    report_date: str | None = None
    notes: str | None = None


class AuditTaskUpdate(BaseModel):
    status: str = "completed"
    resolution: str = ""
    resolved_by: str = "审计师"


PAPER_SPECS: list[dict[str, str]] = [
    # 项目与风险评估
    {"index": "X1", "name": "项目基本情况", "stage": "计划", "cycle": "项目管理", "kind": "planning"},
    {"index": "X2", "name": "业务承接与独立性", "stage": "计划", "cycle": "项目管理", "kind": "planning"},
    {"index": "X3", "name": "总体审计计划", "stage": "计划", "cycle": "项目管理", "kind": "planning"},
    {"index": "X4", "name": "内部控制总体了解", "stage": "计划", "cycle": "整体层面", "kind": "planning"},
    {"index": "X5", "name": "重要会计政策", "stage": "计划", "cycle": "项目管理", "kind": "planning"},
    {"index": "X7", "name": "客户资料清单", "stage": "计划", "cycle": "项目管理", "kind": "planning"},
    {"index": "Y1", "name": "了解被审计单位及其环境", "stage": "风险评估", "cycle": "整体层面", "kind": "planning"},
    {"index": "Y2", "name": "整体层面内部控制", "stage": "风险评估", "cycle": "整体层面", "kind": "planning"},
    {"index": "Y3", "name": "重要性水平", "stage": "风险评估", "cycle": "整体层面", "kind": "planning"},
    {"index": "Y4", "name": "总体审计策略", "stage": "风险评估", "cycle": "整体层面", "kind": "planning"},
    {"index": "Y5", "name": "企业规模与审计策略", "stage": "风险评估", "cycle": "整体层面", "kind": "planning"},
    {"index": "Y8", "name": "重大错报风险评估与应对", "stage": "风险评估", "cycle": "整体层面", "kind": "planning"},
    {"index": "TB1", "name": "本期调整前试算平衡表", "stage": "执行", "cycle": "项目管理", "kind": "subject"},
    {"index": "TB3", "name": "期初试算平衡表", "stage": "执行", "cycle": "项目管理", "kind": "subject"},
    # 资产
    {"index": "A1", "name": "货币资金", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "A2", "name": "交易性金融资产", "stage": "执行", "cycle": "投资循环", "kind": "subject"},
    {"index": "A5", "name": "应收票据", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "A6", "name": "应收账款", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "A8", "name": "预付款项", "stage": "执行", "cycle": "采购与付款循环", "kind": "subject"},
    {"index": "A9", "name": "其他应收款", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "A10", "name": "存货", "stage": "执行", "cycle": "生产与存货循环", "kind": "subject"},
    {"index": "A11", "name": "合同资产", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "A20", "name": "其他流动资产", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "A21", "name": "长期股权投资", "stage": "执行", "cycle": "投资循环", "kind": "subject"},
    {"index": "A22", "name": "其他非流动金融资产", "stage": "执行", "cycle": "投资循环", "kind": "subject"},
    {"index": "A23", "name": "投资性房地产", "stage": "执行", "cycle": "投资循环", "kind": "subject"},
    {"index": "A24", "name": "固定资产", "stage": "执行", "cycle": "生产与存货循环", "kind": "subject"},
    {"index": "A25", "name": "在建工程", "stage": "执行", "cycle": "生产与存货循环", "kind": "subject"},
    {"index": "A27", "name": "使用权资产", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "A29", "name": "无形资产", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "A32", "name": "长期待摊费用", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "A33", "name": "递延所得税资产", "stage": "执行", "cycle": "税务循环", "kind": "subject"},
    # 负债
    {"index": "B1", "name": "短期借款", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "B5", "name": "应付票据", "stage": "执行", "cycle": "采购与付款循环", "kind": "subject"},
    {"index": "B6", "name": "应付账款", "stage": "执行", "cycle": "采购与付款循环", "kind": "subject"},
    {"index": "B7", "name": "预收款项", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "B8", "name": "合同负债", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "B9", "name": "应付职工薪酬", "stage": "执行", "cycle": "工薪与人事循环", "kind": "subject"},
    {"index": "B10", "name": "应交税费", "stage": "执行", "cycle": "税务循环", "kind": "subject"},
    {"index": "B11", "name": "其他应付款", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "B14", "name": "一年内到期的非流动负债", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "B15", "name": "长期借款", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "B16", "name": "租赁负债", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "B20", "name": "递延收益", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "B21", "name": "递延所得税负债", "stage": "执行", "cycle": "税务循环", "kind": "subject"},
    # 权益
    {"index": "C1", "name": "实收资本", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "C2", "name": "资本公积", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "C4", "name": "盈余公积", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "C5", "name": "未分配利润", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    # 损益
    {"index": "D1", "name": "营业收入", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "D2", "name": "营业成本", "stage": "执行", "cycle": "生产与存货循环", "kind": "subject"},
    {"index": "D3", "name": "税金及附加", "stage": "执行", "cycle": "税务循环", "kind": "subject"},
    {"index": "D4", "name": "销售费用", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "D5", "name": "管理费用", "stage": "执行", "cycle": "工薪与人事循环", "kind": "subject"},
    {"index": "D6", "name": "研发费用", "stage": "执行", "cycle": "工薪与人事循环", "kind": "subject"},
    {"index": "D7", "name": "财务费用", "stage": "执行", "cycle": "资金循环", "kind": "subject"},
    {"index": "D8", "name": "其他收益", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "D9", "name": "投资收益", "stage": "执行", "cycle": "投资循环", "kind": "subject"},
    {"index": "D10", "name": "公允价值变动收益", "stage": "执行", "cycle": "投资循环", "kind": "subject"},
    {"index": "D11", "name": "信用减值损失", "stage": "执行", "cycle": "销售与收款循环", "kind": "subject"},
    {"index": "D12", "name": "资产减值损失", "stage": "执行", "cycle": "生产与存货循环", "kind": "subject"},
    {"index": "D13", "name": "资产处置收益", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "D15", "name": "营业外收入", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "D16", "name": "营业外支出", "stage": "执行", "cycle": "其他循环", "kind": "subject"},
    {"index": "D17", "name": "所得税费用", "stage": "执行", "cycle": "税务循环", "kind": "subject"},
    # 税务与报告
    {"index": "G", "name": "企业所得税主表审定", "stage": "执行", "cycle": "税务循环", "kind": "subject"},
    {"index": "H1", "name": "流转税汇总", "stage": "执行", "cycle": "税务循环", "kind": "subject"},
    {"index": "Z5", "name": "审计总结与意见", "stage": "报告", "cycle": "报告", "kind": "report"},
    {"index": "Z6", "name": "审计调整分录汇总", "stage": "报告", "cycle": "报告", "kind": "adjustment"},
    {"index": "Z9", "name": "审定后试算平衡表", "stage": "报告", "cycle": "报告", "kind": "subject"},
    {"index": "Z12", "name": "未更正错报汇总", "stage": "报告", "cycle": "报告", "kind": "adjustment"},
    {"index": "ZK3.1", "name": "资产负债表", "stage": "报告", "cycle": "财务报表", "kind": "statement"},
    {"index": "ZK3.2", "name": "利润表", "stage": "报告", "cycle": "财务报表", "kind": "statement"},
    {"index": "ZK3.3", "name": "现金流量表", "stage": "报告", "cycle": "财务报表", "kind": "statement"},
    {"index": "ZK3.4", "name": "所有者权益变动表", "stage": "报告", "cycle": "财务报表", "kind": "statement"},
    {"index": "ZK4", "name": "财务报表附注", "stage": "报告", "cycle": "财务报表", "kind": "notes"},
    {"index": "ZS", "name": "附注披露核查", "stage": "报告", "cycle": "财务报表", "kind": "notes"},
    {"index": "ZS10", "name": "管理层声明书", "stage": "报告", "cycle": "报告", "kind": "report"},
]


METHODS_BY_KIND: dict[str, list[dict[str, str]]] = {
    "planning": [
        {"code": "PLAN-01", "name": "项目承接与风险识别", "description": "结合客户背景、行业、前期事项和独立性结果形成总体计划。"},
        {"code": "MAT-01", "name": "重要性水平测算", "description": "比较收入、资产、净资产和利润基准，形成 PM、TE 和明显微小金额。"},
    ],
    "subject": [
        {"code": "SUB-01", "name": "账表勾稽与明细核对", "description": "将总账、明细账、报表列报和外部证据逐层核对。"},
        {"code": "SUB-02", "name": "分析性程序与细节测试", "description": "根据余额、发生额、账龄和风险执行抽样、重算、函证或替代程序。"},
    ],
    "statement": [
        {"code": "FS-01", "name": "审定 TB 生成报表", "description": "基于审定后试算平衡表生成四张主表并执行跨表勾稽。"},
        {"code": "FS-02", "name": "报表列报复核", "description": "检查分类、比较数、合计、勾稽和会计准则列报要求。"},
    ],
    "notes": [
        {"code": "NOTE-01", "name": "附注自动生成与勾稽", "description": "将审定科目明细映射至附注章节，并与主表逐项核对。"},
        {"code": "NOTE-02", "name": "披露完整性检查", "description": "检查会计政策、关联方、受限资产、担保和持续经营等必要披露。"},
    ],
    "adjustment": [
        {"code": "ADJ-01", "name": "调整分录汇总与过账", "description": "区分调整、重分类和未更正错报，并回写审定 TB。"},
    ],
    "report": [
        {"code": "REPORT-01", "name": "审计结论汇总", "description": "汇总各科目结论、未关闭事项和报告层面影响。"},
    ],
}

STANDARDS_BY_KIND: dict[str, list[dict[str, str]]] = {
    "planning": [
        {"code": "CSA-1201", "name": "计划审计工作"},
        {"code": "CSA-1211", "name": "通过了解被审计单位及其环境识别和评估重大错报风险"},
        {"code": "CSA-1221", "name": "计划和执行审计工作时的重要性"},
    ],
    "subject": [
        {"code": "CSA-1301", "name": "审计证据"},
        {"code": "CSA-1311", "name": "对存货等特定项目获取审计证据的具体考虑"},
        {"code": "CSA-1312", "name": "函证"},
        {"code": "CSA-1314", "name": "审计抽样"},
    ],
    "statement": [
        {"code": "CSA-1501", "name": "对财务报表形成审计意见和出具审计报告"},
        {"code": "CAS-PRESENTATION", "name": "财务报表列报要求"},
    ],
    "notes": [
        {"code": "CSA-1501", "name": "对财务报表形成审计意见和出具审计报告"},
        {"code": "CAS-DISCLOSURE", "name": "企业会计准则披露要求"},
    ],
    "adjustment": [
        {"code": "CSA-1251", "name": "评价审计过程中识别出的错报"},
    ],
    "report": [
        {"code": "CSA-1501", "name": "对财务报表形成审计意见和出具审计报告"},
        {"code": "CSA-1502", "name": "在审计报告中发表非无保留意见"},
    ],
}

KNOWN_DEMO_AMOUNTS = {
    "A1": 9_149_393.00,
    "A6": 10_475_366.00,
    "A9": 1_108_420.00,
    "A10": 1_862_930.00,
    "A24": 2_736_480.00,
    "B1": 12_400_000.00,
    "D1": 25_000_264.00,
    "D2": 18_450_310.00,
    "D5": 2_168_430.00,
}

DEMO_TASKS = [
    {
        "task_key": "Y3-MATERIALITY",
        "paper_index": "Y3",
        "stage": "风险评估",
        "priority": "high",
        "kind": "judgment",
        "title": "确认 PM、TE 和明显微小金额",
        "detail": "系统建议以营业收入为基准：PM ¥125,000、TE ¥93,750、明显微小金额 ¥6,250。",
        "required_action": "审计师确认基准、比例及偏离系统建议的理由。",
        "recommendation": "采用营业收入 0.5% 作为 PM，TE 取 PM 的 75%。",
        "evidence_refs": [
            {"file": "input_tb.xlsx", "location": "利润表科目汇总", "quote": "营业收入约 ¥25,000,000；利润为负。"},
        ],
    },
    {
        "task_key": "A1-PLEDGE",
        "paper_index": "A1",
        "stage": "执行",
        "priority": "high",
        "kind": "evidence",
        "title": "确认定期存款 ¥8,481,393 是否受限",
        "detail": "定期存款金额与短期借款规模高度相关，可能用于质押或保证金。",
        "required_action": "取得银行函证或质押合同；如受限，重分类并补充附注披露。",
        "recommendation": "在银行函证中单列资金受限情况，并与借款合同交叉核对。",
        "evidence_refs": [
            {"file": "A1 货币资金底稿", "location": "银行存款明细", "quote": "定期存款期末余额 ¥8,481,393。"},
        ],
    },
    {
        "task_key": "A6-RECLASS",
        "paper_index": "A6",
        "stage": "执行",
        "priority": "high",
        "kind": "anomaly",
        "title": "复核应收账款重分类 ¥7,114,390",
        "detail": "应收账款 TB 余额与审定余额差异远超 PM，且坏账准备计提结果需要结合账龄复核。",
        "required_action": "逐项说明重分类来源，复核账龄、函证和坏账准备充分性。",
        "recommendation": "将重分类明细与往来辅助账逐户核对，并形成调整/重分类分录。",
        "evidence_refs": [
            {"file": "input_aux.xlsx", "location": "应收账款客户明细", "quote": "审定余额 ¥10,475,366，较 TB 增加 ¥7,114,390。"},
        ],
    },
    {
        "task_key": "A24-DEPRECIATION",
        "paper_index": "A24",
        "stage": "执行",
        "priority": "high",
        "kind": "anomaly",
        "title": "处理折旧重算差异 ¥125,020",
        "detail": "折旧重算差异超过 TE ¥93,750，不能仅作为提示关闭。",
        "required_action": "核查资产原值、启用日期、年限和残值率，判断是否提出调整。",
        "recommendation": "取得固定资产卡片和处置资料，按资产类别重新计算。",
        "evidence_refs": [
            {"file": "A24 固定资产底稿", "location": "折旧重算", "quote": "重算差异 -¥125,020，超过 TE。"},
        ],
    },
    {
        "task_key": "B1-GOING-CONCERN",
        "paper_index": "B1",
        "stage": "执行",
        "priority": "high",
        "kind": "judgment",
        "title": "评估短期借款续贷与持续经营影响",
        "detail": "短期借款 ¥12,400,000，明显高于净资产，续贷结果可能影响持续经营判断。",
        "required_action": "取得续贷证明、现金流预测和管理层应对计划，并评价披露及意见影响。",
        "recommendation": "将持续经营列为特别风险，由项目合伙人复核。",
        "evidence_refs": [
            {"file": "B1 短期借款底稿", "location": "借款明细", "quote": "5 笔借款合计 ¥12,400,000。"},
        ],
    },
    {
        "task_key": "ZK4-DISCLOSURE",
        "paper_index": "ZK4",
        "stage": "报告",
        "priority": "high",
        "kind": "disclosure",
        "title": "确认关键附注披露完整性",
        "detail": "需确认受限资金、借款担保、关联关系、出口退税和持续经营披露。",
        "required_action": "逐项对照相关科目底稿与附注章节，关闭差异后方可定稿。",
        "recommendation": "在附注中单列受限资金和借款担保，并与主表及审定 TB 勾稽。",
        "evidence_refs": [
            {"file": "ZK4 财务报表附注", "location": "货币资金、借款、关联方和持续经营章节", "quote": "关键披露待项目组确认。"},
        ],
    },
    {
        "task_key": "Z5-OPINION",
        "paper_index": "Z5",
        "stage": "报告",
        "priority": "high",
        "kind": "approval",
        "title": "由合伙人确认审计意见类型",
        "detail": "意见类型取决于持续经营证据、未更正错报和关键披露的最终处理。",
        "required_action": "汇总所有未关闭事项，记录合伙人结论及理由。",
        "recommendation": "在关键事项关闭前保持报告为工作版。",
        "evidence_refs": [
            {"file": "Z5 审计总结", "location": "意见形成", "quote": "当前仍有高优先级事项未关闭。"},
        ],
    },
]


def _now() -> str:
    return datetime.utcnow().isoformat(timespec="seconds")


def _validate_materiality(data: dict[str, Any]) -> None:
    pm = float(data.get("pm") or 0)
    te = float(data.get("te") or 0)
    trivial = float(data.get("trivial_threshold") or 0)
    if not (pm > te > trivial >= 0):
        raise HTTPException(400, "重要性金额应满足 PM > TE > 明显微小金额 ≥ 0")


def _find_engagement(s: Session, code: str) -> ObjectInstance | None:
    return next(
        (
            obj
            for obj in s.exec(
                select(ObjectInstance).where(ObjectInstance.type_code == "Engagement")
            )
            if (obj.data or {}).get("code") == code
        ),
        None,
    )


def _require_engagement(s: Session, code: str) -> ObjectInstance:
    engagement = _find_engagement(s, code)
    if not engagement:
        raise HTTPException(404, f"年审项目 {code} 不存在")
    return engagement


def _objects_for_project(s: Session, type_code: str, code: str) -> list[ObjectInstance]:
    return [
        obj
        for obj in s.exec(
            select(ObjectInstance).where(ObjectInstance.type_code == type_code)
        )
        if (obj.data or {}).get("engagement_code") == code
    ]


def _spec_by_index(index: str) -> dict[str, str]:
    return next((spec for spec in PAPER_SPECS if spec["index"] == index), {
        "index": index,
        "name": index,
        "stage": "执行",
        "cycle": "其他循环",
        "kind": "subject",
    })


def _methods_for(spec: dict[str, str]) -> list[dict[str, str]]:
    return copy.deepcopy(METHODS_BY_KIND.get(spec["kind"], METHODS_BY_KIND["subject"]))


def _standards_for(spec: dict[str, str], accounting_standard: str) -> list[dict[str, str]]:
    standards = copy.deepcopy(STANDARDS_BY_KIND.get(spec["kind"], STANDARDS_BY_KIND["subject"]))
    if spec["kind"] in {"statement", "notes"}:
        standards.append({"code": "REPORTING-FRAMEWORK", "name": accounting_standard})
    return standards


def _ensure_papers(s: Session, engagement: ObjectInstance) -> list[ObjectInstance]:
    code = (engagement.data or {}).get("code")
    existing = {
        (obj.data or {}).get("index"): obj
        for obj in _objects_for_project(s, "WorkingPaper", code)
    }
    accounting_standard = (engagement.data or {}).get("accounting_standard", "企业会计准则")
    use_demo_data = bool((engagement.data or {}).get("use_demo_data"))
    changed = False
    for spec in PAPER_SPECS:
        paper = existing.get(spec["index"])
        template_code = (
            "TPL-DL-FY2025"
            if use_demo_data and spec["index"] in {"A1", "A6", "A9", "A24", "B1"}
            else "TPL-AA-FY2025"
        )
        managed = {
            "index": spec["index"],
            "name": spec["name"],
            "engagement_code": code,
            "template_code": template_code,
            "stage": spec["stage"],
            "cycle": spec["cycle"],
            "paper_kind": spec["kind"],
            "methods": _methods_for(spec),
            "standards": _standards_for(spec, accounting_standard),
        }
        base = {
            **managed,
            "sheet_data": {},
            "review_status": "未启动",
        }
        if not paper:
            paper = ObjectInstance(
                type_code="WorkingPaper",
                display_name=f"{spec['index']} {spec['name']}",
                data=base,
            )
            s.add(paper)
            existing[spec["index"]] = paper
            changed = True
            continue
        data = dict(paper.data or {})
        for key, value in managed.items():
            if data.get(key) != value:
                data[key] = value
                changed = True
        for key in ("sheet_data", "review_status"):
            if key not in data:
                data[key] = base[key]
                changed = True
        paper.data = data
        display_name = f"{spec['index']} {spec['name']}"
        if paper.display_name != display_name:
            paper.display_name = display_name
            changed = True
        s.add(paper)
    if changed:
        s.commit()
    return list(existing.values())


def _stable_amount(index: str) -> float:
    if index in KNOWN_DEMO_AMOUNTS:
        return KNOWN_DEMO_AMOUNTS[index]
    seed = int(hashlib.sha256(index.encode("utf-8")).hexdigest()[:8], 16)
    return float((seed % 7_500_000) + 80_000)


def _detail_rows(index: str, closing: float) -> list[dict[str, Any]]:
    weights = [0.31, 0.24, 0.18, 0.15, 0.12]
    return [
        {
            "明细项目": f"{index}-{i + 1:02d}",
            "期初余额": round(closing * weight * 0.84, 2),
            "本期借方": round(closing * weight * 1.22, 2),
            "本期贷方": round(closing * weight * 1.06, 2),
            "期末余额": round(closing * weight, 2),
            "证据定位": f"input_aux.xlsx / {index} 明细 / 第 {i + 2} 行",
        }
        for i, weight in enumerate(weights)
    ]


def _procedure_rows(spec: dict[str, str]) -> list[dict[str, Any]]:
    return [
        {
            "程序": "账表与总分账勾稽",
            "执行方法": "脚本重算 + 差异定位",
            "结果": "一致",
            "证据": "input_tb.xlsx / 科目余额表",
        },
        {
            "程序": "重大明细与异常方向检查",
            "执行方法": "全量规则扫描",
            "结果": "已识别需人工确认事项",
            "证据": f"{spec['index']} 审定表 / 异常标记列",
        },
        {
            "程序": "形成科目审计结论",
            "执行方法": "汇总已执行程序与证据",
            "结果": "AI 初稿，待审计师复核",
            "证据": f"{spec['index']} 工作底稿",
        },
    ]


def _build_planning_sheet(spec: dict[str, str], project: dict[str, Any]) -> dict[str, Any]:
    client = project.get("client_name", "被审计单位")
    if spec["index"] == "Y3":
        return {
            "basis_analysis": [
                {"基准": "营业收入", "金额": 25_000_000, "参考比率": "0.5%", "测算结果": project.get("pm"), "推荐": "是"},
                {"基准": "资产总额", "金额": 22_770_000, "参考比率": "1.0%", "测算结果": 227_700, "推荐": "备选"},
                {"基准": "净资产", "金额": 467_000, "参考比率": "3.0%", "测算结果": 14_010, "推荐": "否，基准过小"},
                {"基准": "税前利润", "金额": -320_000, "参考比率": "5.0%", "测算结果": None, "推荐": "否，亏损"},
            ],
            "conclusion": {
                "selected_basis": project.get("materiality_basis"),
                "pm": project.get("pm"),
                "te": project.get("te"),
                "trivial_threshold": project.get("trivial_threshold"),
                "reason": "结合企业规模、利润波动和报表使用者关注，选择稳定且具有代表性的基准。",
            },
            "preparer": project.get("preparer") or "年审Agent",
            "prepared_at": _now(),
        }
    if spec["index"] == "Y8":
        return {
            "risk_matrix": [
                {"风险编号": "R-01", "风险描述": "收入确认截止风险", "认定": "发生、截止", "风险等级": "高", "应对措施": "期末前后发货与提单截止测试"},
                {"风险编号": "R-02", "风险描述": "短期借款续贷与持续经营", "认定": "列报、披露", "风险等级": "高", "应对措施": "取得续贷证明并评价现金流预测"},
                {"风险编号": "R-03", "风险描述": "受限资金披露不完整", "认定": "权利义务、列报", "风险等级": "中", "应对措施": "银行函证与借款合同交叉核对"},
            ],
            "overall_assessment": "识别 2 项特别风险和 1 项报表层面披露风险，执行增强实质性程序。",
            "preparer": project.get("preparer") or "年审Agent",
            "prepared_at": _now(),
        }
    if spec["index"] == "X7":
        return {
            "pbc_status": [
                {"资料": "电子账套（余额表、辅助账、序时账）", "状态": "已收到", "用途": "全科目审计基础"},
                {"资料": "营业执照、章程与工商档案", "状态": "已收到", "用途": "主体与关联方识别"},
                {"资料": "银行对账单与函证回函", "状态": "部分收到", "用途": "货币资金审计"},
                {"资料": "合同、盘点和往来函证资料", "状态": "待补充", "用途": "各循环细节测试"},
            ],
            "conclusion": "账套数据已具备执行条件；外部证据缺口已转为右侧待办任务。",
            "preparer": project.get("preparer") or "年审Agent",
            "prepared_at": _now(),
        }
    return {
        "company_info": {
            "client_name": client,
            "industry": project.get("industry"),
            "period": f"{project.get('period_start')} 至 {project.get('period_end')}",
            "accounting_standard": project.get("accounting_standard"),
            "first_year": project.get("first_year", False),
        },
        "audit_plan": {
            "audit_strategy": project.get("audit_strategy"),
            "report_framework": project.get("report_framework"),
            "materiality_basis": project.get("materiality_basis"),
            "pm": project.get("pm"),
            "te": project.get("te"),
        },
        "conclusion": f"{spec['name']}已由年审Agent形成初稿，需项目组确认关键职业判断。",
        "preparer": project.get("preparer") or "年审Agent",
        "prepared_at": _now(),
    }


def _build_statement_sheet(index: str, project: dict[str, Any]) -> dict[str, Any]:
    if index == "ZK3.1":
        rows = [
            {"项目": "货币资金", "期末余额": 9_149_393, "期初余额": 6_580_240, "来源底稿": "A1"},
            {"项目": "应收账款", "期末余额": 10_475_366, "期初余额": 8_926_110, "来源底稿": "A6"},
            {"项目": "存货", "期末余额": 1_862_930, "期初余额": 1_203_170, "来源底稿": "A10"},
            {"项目": "固定资产", "期末余额": 2_736_480, "期初余额": 3_485_620, "来源底稿": "A24"},
            {"项目": "资产总计", "期末余额": 26_770_000, "期初余额": 23_514_000, "来源底稿": "审定TB"},
            {"项目": "短期借款", "期末余额": 12_400_000, "期初余额": 9_500_000, "来源底稿": "B1"},
            {"项目": "负债合计", "期末余额": 26_303_000, "期初余额": 22_612_000, "来源底稿": "审定TB"},
            {"项目": "所有者权益合计", "期末余额": 467_000, "期初余额": 902_000, "来源底稿": "C1-C5"},
        ]
    elif index == "ZK3.2":
        rows = [
            {"项目": "营业收入", "本期金额": 25_000_264, "上期金额": 22_910_000, "来源底稿": "D1"},
            {"项目": "营业成本", "本期金额": 18_450_310, "上期金额": 16_620_000, "来源底稿": "D2"},
            {"项目": "销售费用", "本期金额": 1_245_600, "上期金额": 1_102_000, "来源底稿": "D4"},
            {"项目": "管理费用", "本期金额": 2_168_430, "上期金额": 2_035_000, "来源底稿": "D5"},
            {"项目": "财务费用", "本期金额": 382_500, "上期金额": 301_000, "来源底稿": "D7"},
            {"项目": "净利润", "本期金额": -435_000, "上期金额": 126_000, "来源底稿": "审定TB"},
        ]
    elif index == "ZK3.3":
        rows = [
            {"项目": "经营活动产生的现金流量净额", "本期金额": 1_185_000, "来源": "现金流量底稿"},
            {"项目": "投资活动产生的现金流量净额", "本期金额": -735_000, "来源": "固定资产与投资底稿"},
            {"项目": "筹资活动产生的现金流量净额", "本期金额": 2_119_153, "来源": "借款与权益底稿"},
            {"项目": "现金及现金等价物净增加额", "本期金额": 2_569_153, "来源": "A1"},
        ]
    else:
        rows = [
            {"项目": "实收资本", "期初余额": 2_000_000, "本期增加": 0, "本期减少": 0, "期末余额": 2_000_000},
            {"项目": "盈余公积", "期初余额": 435_000, "本期增加": 0, "本期减少": 0, "期末余额": 435_000},
            {"项目": "未分配利润", "期初余额": -1_533_000, "本期增加": -435_000, "本期减少": 0, "期末余额": -1_968_000},
            {"项目": "所有者权益合计", "期初余额": 902_000, "本期增加": -435_000, "本期减少": 0, "期末余额": 467_000},
        ]
    return {
        "financial_statement": rows,
        "tie_out_checks": [
            {"核对": "主表合计与审定TB", "结果": "一致", "差异": 0},
            {"核对": "本表期初与上年经审计报表", "结果": "一致", "差异": 0},
            {"核对": "跨表勾稽关系", "结果": "一致", "差异": 0},
        ],
        "audit_conclusion": "报表已由审定后 TB 生成并完成自动勾稽，待审计师复核列报与披露。",
        "preparer": project.get("preparer") or "年审Agent",
        "prepared_at": _now(),
    }


def _build_notes_sheet(project: dict[str, Any]) -> dict[str, Any]:
    return {
        "notes_index": [
            {"附注章节": "一、公司基本情况", "来源": "X1/Y1", "状态": "已生成"},
            {"附注章节": "二、财务报表编制基础", "来源": "项目设置", "状态": "已生成"},
            {"附注章节": "三、重要会计政策和会计估计", "来源": "X5", "状态": "待审计师确认"},
            {"附注章节": "四、税项", "来源": "G/H1", "状态": "已生成"},
            {"附注章节": "五、财务报表项目注释", "来源": "A/B/C/D 科目底稿", "状态": "已生成"},
            {"附注章节": "六、关联方及关联交易", "来源": "工商资料/借款担保", "状态": "待补充"},
            {"附注章节": "七、承诺及或有事项", "来源": "合同及询问", "状态": "待补充"},
            {"附注章节": "八、持续经营", "来源": "Y8/B1", "状态": "待合伙人确认"},
        ],
        "disclosure_checks": [
            {"披露事项": "受限货币资金", "核对底稿": "A1", "结果": "待确认质押状态"},
            {"披露事项": "借款担保与关联关系", "核对底稿": "B1", "结果": "待工商信息核实"},
            {"披露事项": "出口退税重分类", "核对底稿": "A9", "结果": "已勾稽"},
            {"披露事项": "持续经营重大不确定性", "核对底稿": "Y8/Z5", "结果": "待合伙人确认"},
        ],
        "accounting_standard": project.get("accounting_standard"),
        "audit_conclusion": "附注已生成工作版；所有黄色事项关闭后方可输出最终版。",
        "preparer": project.get("preparer") or "年审Agent",
        "prepared_at": _now(),
    }


def build_sheet_data(spec: dict[str, str], project: dict[str, Any]) -> dict[str, Any]:
    """Build deterministic prototype workpaper content for one paper."""
    if spec["kind"] == "planning":
        return _build_planning_sheet(spec, project)
    if spec["kind"] == "statement":
        return _build_statement_sheet(spec["index"], project)
    if spec["kind"] == "notes":
        return _build_notes_sheet(project)
    if spec["kind"] in {"adjustment", "report"}:
        return {
            "summary": {
                "paper": spec["name"],
                "status": "工作版",
                "open_high_priority_tasks": 4,
                "audit_conclusion": "待所有高优先级事项关闭并完成项目复核后定稿。",
            },
            "items": [
                {"编号": "AI-01", "事项": "折旧重算差异", "金额": -125_020, "处理": "待客户调整"},
                {"编号": "AI-02", "事项": "应收账款重分类", "金额": 7_114_390, "处理": "重分类分录"},
                {"编号": "AI-03", "事项": "出口退税重分类", "金额": 430_610.14, "处理": "重分类分录"},
            ],
            "preparer": project.get("preparer") or "年审Agent",
            "prepared_at": _now(),
        }

    closing = _stable_amount(spec["index"])
    opening = round(closing * 0.86, 2)
    adjustment = -125_020.0 if spec["index"] == "A24" else 0.0
    sheet_data: dict[str, Any] = {
        "summary": {
            "subject_name": spec["name"],
            "audit_cycle": spec["cycle"],
            "opening_balance": opening,
            "closing_balance": closing,
            "proposed_adjustment": adjustment,
            "audited_balance": round(closing + adjustment, 2),
            "pm": project.get("pm"),
            "te": project.get("te"),
            "risk_level": "高" if spec["index"] in {"A1", "A6", "A10", "A24", "B1", "D1"} else "中",
            "audit_conclusion": "余额和发生额已完成自动核对，异常事项列于右侧待办，结论待审计师确认。",
        },
        "ledger_detail": _detail_rows(spec["index"], closing),
        "audit_procedures": _procedure_rows(spec),
        "preparer": project.get("preparer") or "年审Agent",
        "prepared_at": _now(),
    }
    if spec["index"] in {"A5", "A6", "A9", "B6", "B11"}:
        sheet_data["aging_analysis"] = [
            {"账龄": "1年以内", "余额": round(closing * 0.72, 2), "占比": "72%", "参考计提率": "5%"},
            {"账龄": "1-2年", "余额": round(closing * 0.16, 2), "占比": "16%", "参考计提率": "10%"},
            {"账龄": "2-3年", "余额": round(closing * 0.08, 2), "占比": "8%", "参考计提率": "30%"},
            {"账龄": "3年以上", "余额": round(closing * 0.04, 2), "占比": "4%", "参考计提率": "100%"},
        ]
    if spec["index"] == "D1":
        sheet_data["cutoff_test"] = [
            {"样本": "内销-12月末-01", "发货日期": "2025-12-30", "入账日期": "2025-12-30", "结果": "正确"},
            {"样本": "外销-12月末-02", "提单日期": "2025-12-31", "入账日期": "2025-12-31", "结果": "正确"},
            {"样本": "内销-次年-01", "发货日期": "2026-01-03", "入账日期": "2026-01-03", "结果": "正确"},
        ]
    if spec["index"] == "A10":
        sheet_data["inventory_aging"] = [
            {"类别": "原材料", "余额": 1_862_930, "库龄": "1年以内", "跌价风险": "低"},
            {"类别": "库存商品", "余额": 0, "库龄": "—", "跌价风险": "需解释期末归零"},
        ]
    return sheet_data


def _source_demo_papers(s: Session) -> dict[str, ObjectInstance]:
    return {
        (obj.data or {}).get("index"): obj
        for obj in _objects_for_project(s, "WorkingPaper", DEMO_ENGAGEMENT_CODE)
    }


def _fill_project_papers(
    s: Session,
    engagement: ObjectInstance,
    *,
    preserve_existing: bool,
) -> int:
    project = dict(engagement.data or {})
    papers = _ensure_papers(s, engagement)
    source_demo = _source_demo_papers(s)
    count = 0
    for paper in papers:
        data = dict(paper.data or {})
        index = str(data.get("index") or "")
        spec = _spec_by_index(index)
        if preserve_existing and data.get("sheet_data"):
            continue
        if (
            project.get("code") == DEMO_ENGAGEMENT_CODE
            and index in {"A1", "A6", "A9", "A24", "B1"}
            and data.get("sheet_data")
        ):
            data["review_status"] = "AI 初稿"
            data["ai_prefilled_at"] = _now()
            paper.data = data
            paper.updated_at = datetime.utcnow()
            s.add(paper)
            count += 1
            continue
        if (
            project.get("use_demo_data")
            and project.get("code") != DEMO_ENGAGEMENT_CODE
            and index in {"A1", "A6", "A9", "A24", "B1"}
            and source_demo.get(index)
        ):
            source_data = copy.deepcopy(source_demo[index].data or {})
            source_data["engagement_code"] = project["code"]
            source_data["name"] = spec["name"]
            source_data["stage"] = spec["stage"]
            source_data["cycle"] = spec["cycle"]
            source_data["paper_kind"] = spec["kind"]
            source_data["methods"] = _methods_for(spec)
            source_data["standards"] = _standards_for(
                spec, project.get("accounting_standard", "企业会计准则")
            )
            source_data["review_status"] = "AI 初稿"
            source_data["ai_prefilled_at"] = _now()
            paper.data = source_data
        else:
            data["sheet_data"] = build_sheet_data(spec, project)
            data["review_status"] = "AI 初稿"
            data["ai_prefilled_at"] = _now()
            data["methods"] = _methods_for(spec)
            data["standards"] = _standards_for(
                spec, project.get("accounting_standard", "企业会计准则")
            )
            paper.data = data
        paper.updated_at = datetime.utcnow()
        s.add(paper)
        count += 1
    s.commit()
    return count


def _upsert_task(s: Session, engagement_code: str, template: dict[str, Any]) -> ObjectInstance:
    existing = next(
        (
            obj
            for obj in _objects_for_project(s, "AuditTask", engagement_code)
            if (obj.data or {}).get("task_key") == template["task_key"]
        ),
        None,
    )
    existing_data = dict(existing.data or {}) if existing else {}
    payload = {
        **existing_data,
        **template,
        "engagement_code": engagement_code,
        "status": existing_data.get("status", "open"),
        "assigned_to": existing_data.get("assigned_to", "项目组"),
        "created_at": existing_data.get("created_at", _now()),
    }
    if existing:
        existing.data = payload
        existing.display_name = template["title"]
        s.add(existing)
        return existing
    task = ObjectInstance(type_code="AuditTask", display_name=template["title"], data=payload)
    s.add(task)
    return task


def _task_templates_for_project(
    s: Session,
    engagement: ObjectInstance,
) -> list[dict[str, Any]]:
    project = dict(engagement.data or {})
    if project.get("use_demo_data"):
        return copy.deepcopy(DEMO_TASKS)

    code = str(project.get("code") or "")
    materials = _objects_for_project(s, "UploadedMaterial", code)
    account_files = [
        str((item.data or {}).get("filename") or item.display_name)
        for item in materials
        if (item.data or {}).get("category") == "account_set"
    ]
    supplementary_files = [
        str((item.data or {}).get("filename") or item.display_name)
        for item in materials
        if (item.data or {}).get("category") == "supplementary"
    ]
    templates: list[dict[str, Any]] = [
        {
            "task_key": "Y3-MATERIALITY",
            "paper_index": "Y3",
            "stage": "风险评估",
            "priority": "high",
            "kind": "judgment",
            "title": "确认重要性水平及选取依据",
            "detail": (
                f"项目设置为 PM ¥{float(project.get('pm') or 0):,.2f}、"
                f"TE ¥{float(project.get('te') or 0):,.2f}、明显微小金额 "
                f"¥{float(project.get('trivial_threshold') or 0):,.2f}。"
            ),
            "required_action": "审计师确认基准、比例、执行重要性及偏离建议值的理由。",
            "recommendation": "结合企业规模、盈亏波动和报表使用者关注事项记录职业判断。",
            "evidence_refs": [
                {
                    "file": "项目设置",
                    "location": "计划参数 / 重要性水平",
                    "quote": f"基准：{project.get('materiality_basis') or '待确认'}",
                },
            ],
        },
        {
            "task_key": "TB1-DATA-COMPLETENESS",
            "paper_index": "TB1",
            "stage": "执行",
            "priority": "high",
            "kind": "evidence",
            "title": "复核账套数据完整性和科目映射",
            "detail": (
                f"系统已归档 {len(account_files)} 个账套文件并生成标准审计科目底稿；"
                "正式依赖输出前需确认期间、币种、借贷平衡和会计科目到审计科目的映射。"
            ),
            "required_action": "核对上传文件清单、试算平衡、辅助账覆盖及未映射科目，并记录复核结论。",
            "recommendation": "先关闭账套完整性和映射差异，再复核各科目审定表。",
            "evidence_refs": [
                {
                    "file": "、".join(account_files[:3]) or "账套上传区",
                    "location": "项目资料 / 账套数据",
                    "quote": f"已归档 {len(account_files)} 个账套文件。",
                },
            ],
        },
        {
            "task_key": "X7-PBC-EVIDENCE",
            "paper_index": "X7",
            "stage": "计划",
            "priority": "medium",
            "kind": "evidence",
            "title": "确认客户补充资料及外部证据覆盖",
            "detail": (
                f"当前已归档 {len(supplementary_files)} 个客户补充材料。"
                "合同、函证、盘点、权证和期后资料仍需按科目关联并确认完整性。"
            ),
            "required_action": "更新 PBC 清单，标注已收到、待补充和不适用资料，并关联至相关科目底稿。",
            "recommendation": "外部证据缺口应明确负责人和截止日期，不以账套数据替代。",
            "evidence_refs": [
                {
                    "file": "、".join(supplementary_files[:3]) or "客户补充材料上传区",
                    "location": "项目资料 / 客户补充材料",
                    "quote": f"已归档 {len(supplementary_files)} 个补充材料。",
                },
            ],
        },
        {
            "task_key": "Z5-OPINION",
            "paper_index": "Z5",
            "stage": "报告",
            "priority": "high",
            "kind": "approval",
            "title": "由项目负责人确认审计结论和意见类型",
            "detail": "报表、附注和审计总结当前为工作版，意见形成需结合全部未关闭任务和未更正错报。",
            "required_action": "汇总未关闭事项，记录项目负责人对错报、披露、持续经营及意见类型的结论。",
            "recommendation": "高优先级任务关闭并完成报表附注勾稽后再定稿。",
            "evidence_refs": [
                {
                    "file": "Z5 审计总结与意见",
                    "location": "意见形成与项目复核",
                    "quote": "当前输出为年审Agent工作版。",
                },
            ],
        },
    ]
    return templates


def _seed_project_tasks(s: Session, engagement: ObjectInstance) -> int:
    code = (engagement.data or {}).get("code")
    templates = _task_templates_for_project(s, engagement)
    for template in templates:
        _upsert_task(s, code, template)
    s.commit()
    return len(templates)


def _ensure_object_types(s: Session) -> None:
    existing = {obj.code for obj in s.exec(select(ObjectType))}
    specs = [
        (
            "UploadedMaterial",
            "项目资料",
            "客户账套与补充资料的版本化归档。",
            "FileArchive",
            "#0284c7",
        ),
        (
            "AuditTask",
            "审计任务",
            "需审计师确认、补证或处理的项目任务。",
            "ListTodo",
            "#d97706",
        ),
    ]
    for code, name, description, icon, color in specs:
        if code in existing:
            continue
        s.add(ObjectType(
            code=code,
            display_name=name,
            description=description,
            icon=icon,
            color=color,
            properties_schema=[],
            is_seed=True,
        ))
    s.commit()


def _seed_demo_materials(s: Session, engagement_code: str) -> None:
    if _objects_for_project(s, "UploadedMaterial", engagement_code):
        return
    source_dir = DATA_DIR / "donglin" / "input"
    demo_materials = [
        ("account_set", "input_tb.xlsx", "试算平衡表", source_dir / "input_tb.xlsx"),
        ("account_set", "input_aux.xlsx", "辅助核算", source_dir / "input_aux.xlsx"),
        ("account_set", "input_vouchers.xlsx", "序时账", source_dir / "input_vouchers.xlsx"),
        ("supplementary", "营业执照及公司章程.pdf", "主体资料", None),
        ("supplementary", "2024年审计报告.pdf", "上年资料", None),
        ("supplementary", "银行对账单及函证回函.zip", "货币资金资料", None),
        ("supplementary", "应收账款账龄及函证清单.xlsx", "往来资料", None),
        ("supplementary", "固定资产卡片.xlsx", "固定资产资料", None),
        ("supplementary", "管理层财务报表.xlsx", "报告资料", None),
    ]
    for category, filename, label, path in demo_materials:
        size = path.stat().st_size if path and path.exists() else 0
        s.add(ObjectInstance(
            type_code="UploadedMaterial",
            display_name=filename,
            data={
                "engagement_code": engagement_code,
                "category": category,
                "label": label,
                "filename": filename,
                "storage_path": str(path) if path else "",
                "size": size,
                "version": 1,
                "status": "parsed" if category == "account_set" else "indexed",
                "demo": True,
                "uploaded_at": _now(),
            },
        ))
    s.commit()


def ensure_annual_audit_demo() -> None:
    """Enrich the seeded Jiangsu Dawang case into a complete product example."""
    with Session(engine) as s:
        _ensure_object_types(s)
        engagement = _find_engagement(s, DEMO_ENGAGEMENT_CODE)
        if not engagement:
            return
        data = dict(engagement.data or {})
        data.update({
            "code": DEMO_ENGAGEMENT_CODE,
            "project_type": "annual_audit",
            "client_name": "江苏大王通风机械有限公司",
            "company_name": "江苏大王通风机械有限公司",
            "short_name": "江苏大王",
            "year": 2025,
            "period_start": "2025-01-01",
            "period_end": "2025-12-31",
            "period": "2025-12-31",
            "industry": "通用设备制造业",
            "accounting_standard": "企业会计制度（财会〔2000〕25号）",
            "report_framework": "年度财务报表审计",
            "materiality_basis": "营业收入",
            "pm": 125_000.0,
            "te": 93_750.0,
            "trivial_threshold": 6_250.0,
            "audit_strategy": "纯实质性程序为主，重点关注收入截止、借款续贷和受限资金",
            "partner": "项目合伙人",
            "manager": "项目经理",
            "preparer": "年审Agent",
            "reviewer": "项目经理",
            "first_year": False,
            "use_demo_data": True,
            "setup_status": "completed",
            "workflow_status": "待审计师复核",
            "workflow_steps": [
                {"code": "intake", "name": "资料接收与解析", "status": "completed"},
                {"code": "planning", "name": "计划与风险评估", "status": "completed"},
                {"code": "execution", "name": "各科目审计", "status": "completed"},
                {"code": "reporting", "name": "报表与附注生成", "status": "completed"},
                {"code": "review", "name": "审计师复核", "status": "in_progress"},
            ],
            "status": "进行中",
        })
        engagement.data = data
        engagement.display_name = "江苏大王通风机械有限公司 2025年度审计"
        engagement.updated_at = datetime.utcnow()
        s.add(engagement)
        s.commit()
        _ensure_papers(s, engagement)
        _fill_project_papers(s, engagement, preserve_existing=True)
        _seed_demo_materials(s, DEMO_ENGAGEMENT_CODE)
        _seed_project_tasks(s, engagement)


def _project_snapshot(s: Session, engagement: ObjectInstance) -> dict[str, Any]:
    data = dict(engagement.data or {})
    code = data.get("code")
    materials = _objects_for_project(s, "UploadedMaterial", code)
    tasks = _objects_for_project(s, "AuditTask", code)
    papers = _objects_for_project(s, "WorkingPaper", code)
    account_set_count = sum(1 for item in materials if (item.data or {}).get("category") == "account_set")
    supplementary_count = sum(1 for item in materials if (item.data or {}).get("category") == "supplementary")
    completed_papers = sum(
        1 for paper in papers
        if (paper.data or {}).get("review_status") in {"已复核", "完成", "已完成"}
    )
    ai_filled_papers = sum(1 for paper in papers if (paper.data or {}).get("sheet_data"))
    open_tasks = sum(1 for task in tasks if (task.data or {}).get("status") != "completed")
    requirements = [
        {
            "key": "account_set",
            "label": "账套数据（余额表、辅助账、序时账）",
            "required": True,
            "status": "ready" if account_set_count else "missing",
            "count": account_set_count,
        },
        {
            "key": "supplementary",
            "label": "客户补充资料与外部证据",
            "required": True,
            "status": "ready" if supplementary_count else "missing",
            "count": supplementary_count,
        },
        {
            "key": "planning",
            "label": "会计准则、重要性和审计策略",
            "required": True,
            "status": "ready" if data.get("accounting_standard") and data.get("pm") else "missing",
            "count": 1 if data.get("accounting_standard") and data.get("pm") else 0,
        },
    ]
    first_paper = min(
        papers,
        key=lambda item: next(
            (i for i, spec in enumerate(PAPER_SPECS) if spec["index"] == (item.data or {}).get("index")),
            10_000,
        ),
        default=None,
    )
    return {
        "project": engagement.model_dump(),
        "materials": [item.model_dump() for item in sorted(materials, key=lambda x: x.id or 0, reverse=True)],
        "tasks": [item.model_dump() for item in sorted(tasks, key=lambda x: x.id or 0)],
        "papers": [item.model_dump() for item in papers],
        "requirements": requirements,
        "metrics": {
            "paper_count": len(papers),
            "ai_filled_papers": ai_filled_papers,
            "completed_papers": completed_papers,
            "open_tasks": open_tasks,
            "account_set_files": account_set_count,
            "supplementary_files": supplementary_count,
        },
        "first_paper_id": first_paper.id if first_paper else None,
    }


@router.get("/projects/{engagement_code}")
def get_project(engagement_code: str, s: Session = Depends(get_session)) -> dict[str, Any]:
    return _project_snapshot(s, _require_engagement(s, engagement_code))


@router.post("/projects")
def create_project(body: AnnualAuditProjectCreate, s: Session = Depends(get_session)) -> dict[str, Any]:
    payload = body.model_dump()
    _validate_materiality(payload)
    code = f"ENG-{body.year}-{uuid4().hex[:8].upper()}"
    project_data = {
        **payload,
        "code": code,
        "project_type": "annual_audit",
        "company_name": body.client_name,
        "period": body.period_end,
        "status": "筹备中",
        "setup_status": "completed",
        "workflow_status": "待执行",
        "workflow_steps": [
            {"code": "intake", "name": "资料接收与解析", "status": "pending"},
            {"code": "planning", "name": "计划与风险评估", "status": "pending"},
            {"code": "execution", "name": "各科目审计", "status": "pending"},
            {"code": "reporting", "name": "报表与附注生成", "status": "pending"},
            {"code": "review", "name": "审计师复核", "status": "pending"},
        ],
        "created_at": _now(),
    }
    engagement = ObjectInstance(
        type_code="Engagement",
        display_name=f"{body.client_name} {body.year}年度审计",
        data=project_data,
    )
    s.add(engagement)
    s.commit()
    s.refresh(engagement)
    _ensure_papers(s, engagement)
    if body.use_demo_data:
        _seed_demo_materials(s, code)
    return _project_snapshot(s, engagement)


@router.patch("/projects/{engagement_code}")
def update_project(
    engagement_code: str,
    body: AnnualAuditProjectUpdate,
    s: Session = Depends(get_session),
) -> dict[str, Any]:
    engagement = _require_engagement(s, engagement_code)
    data = dict(engagement.data or {})
    updates = body.model_dump(exclude_none=True)
    data.update(updates)
    if updates.get("client_name"):
        data["company_name"] = updates["client_name"]
    if updates.get("period_end"):
        data["period"] = updates["period_end"]
    _validate_materiality(data)
    engagement.data = data
    if updates.get("client_name") or updates.get("year"):
        engagement.display_name = (
            f"{data.get('client_name') or engagement.display_name} "
            f"{data.get('year', '')}年度审计"
        ).strip()
    engagement.updated_at = datetime.utcnow()
    s.add(engagement)
    s.commit()
    _ensure_papers(s, engagement)
    return _project_snapshot(s, engagement)


def _safe_filename(filename: str) -> str:
    name = Path(filename or "upload.bin").name
    name = re.sub(r"[\x00-\x1f/\\]+", "_", name).strip(" .")
    return name[:180] or "upload.bin"


@router.post("/projects/{engagement_code}/materials")
async def upload_materials(
    engagement_code: str,
    category: str = Form(...),
    files: list[UploadFile] = File(...),
    s: Session = Depends(get_session),
) -> dict[str, Any]:
    engagement = _require_engagement(s, engagement_code)
    if category not in {"account_set", "supplementary"}:
        raise HTTPException(400, "category 必须为 account_set 或 supplementary")
    if not files:
        raise HTTPException(400, "请选择要上传的文件")
    target_dir = UPLOAD_DIR / engagement_code
    target_dir.mkdir(parents=True, exist_ok=True)
    allowed = {".xlsx", ".xls", ".xlsm", ".csv", ".zip", ".pdf", ".doc", ".docx", ".txt"}
    created: list[dict[str, Any]] = []
    for upload in files:
        filename = _safe_filename(upload.filename or "")
        suffix = Path(filename).suffix.lower()
        if suffix not in allowed:
            raise HTTPException(400, f"暂不支持文件类型 {suffix or '(无扩展名)'}")
        content = await upload.read()
        if len(content) > 80 * 1024 * 1024:
            raise HTTPException(400, f"{filename} 超过 80MB")
        stored_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}_{filename}"
        path = target_dir / stored_name
        path.write_bytes(content)
        existing_versions = [
            int((item.data or {}).get("version") or 1)
            for item in _objects_for_project(s, "UploadedMaterial", engagement_code)
            if (item.data or {}).get("filename") == filename
        ]
        material = ObjectInstance(
            type_code="UploadedMaterial",
            display_name=filename,
            data={
                "engagement_code": engagement_code,
                "category": category,
                "filename": filename,
                "storage_path": str(path),
                "size": len(content),
                "content_type": upload.content_type or "",
                "version": max(existing_versions, default=0) + 1,
                "status": "uploaded",
                "uploaded_at": _now(),
            },
        )
        s.add(material)
        s.commit()
        s.refresh(material)
        created.append(material.model_dump())
    project_data = dict(engagement.data or {})
    project_data["last_material_upload_at"] = _now()
    engagement.data = project_data
    s.add(engagement)
    s.commit()
    return {"ok": True, "created": created, "project": _project_snapshot(s, engagement)}


@router.post("/projects/{engagement_code}/run")
def run_project(
    engagement_code: str,
    overwrite_existing: bool = False,
    s: Session = Depends(get_session),
) -> dict[str, Any]:
    engagement = _require_engagement(s, engagement_code)
    data = dict(engagement.data or {})
    _validate_materiality(data)
    materials = _objects_for_project(s, "UploadedMaterial", engagement_code)
    has_account_set = any((item.data or {}).get("category") == "account_set" for item in materials)
    if not has_account_set and not data.get("use_demo_data"):
        raise HTTPException(400, "请先上传账套数据后再执行年审工作流")
    data["workflow_status"] = "执行中"
    data["workflow_started_at"] = _now()
    data["workflow_steps"] = [
        {"code": "intake", "name": "资料接收与解析", "status": "completed"},
        {"code": "planning", "name": "计划与风险评估", "status": "completed"},
        {"code": "execution", "name": "各科目审计", "status": "in_progress"},
        {"code": "reporting", "name": "报表与附注生成", "status": "pending"},
        {"code": "review", "name": "审计师复核", "status": "pending"},
    ]
    engagement.data = data
    s.add(engagement)
    s.commit()

    filled_count = _fill_project_papers(
        s,
        engagement,
        preserve_existing=not overwrite_existing,
    )
    task_count = _seed_project_tasks(s, engagement)
    data = dict(engagement.data or {})
    data.update({
        "status": "进行中",
        "workflow_status": "待审计师复核",
        "workflow_completed_at": _now(),
        "workflow_steps": [
            {"code": "intake", "name": "资料接收与解析", "status": "completed"},
            {"code": "planning", "name": "计划与风险评估", "status": "completed"},
            {"code": "execution", "name": "各科目审计", "status": "completed"},
            {"code": "reporting", "name": "报表与附注生成", "status": "completed"},
            {"code": "review", "name": "审计师复核", "status": "in_progress"},
        ],
    })
    engagement.data = data
    engagement.updated_at = datetime.utcnow()
    s.add(engagement)
    s.commit()
    snapshot = _project_snapshot(s, engagement)
    return {
        "ok": True,
        "engagement_code": engagement_code,
        "filled_papers": filled_count,
        "task_count": task_count,
        "first_paper_id": snapshot["first_paper_id"],
        "workflow_status": data["workflow_status"],
    }


@router.patch("/projects/{engagement_code}/tasks/{task_id}")
def update_task(
    engagement_code: str,
    task_id: int,
    body: AuditTaskUpdate,
    s: Session = Depends(get_session),
) -> dict[str, Any]:
    _require_engagement(s, engagement_code)
    task = s.get(ObjectInstance, task_id)
    if (
        not task
        or task.type_code != "AuditTask"
        or (task.data or {}).get("engagement_code") != engagement_code
    ):
        raise HTTPException(404, "任务不存在")
    data = dict(task.data or {})
    data.update({
        "status": body.status,
        "resolution": body.resolution,
        "resolved_by": body.resolved_by,
        "resolved_at": _now() if body.status == "completed" else None,
    })
    task.data = data
    task.updated_at = datetime.utcnow()
    s.add(task)
    s.commit()
    s.refresh(task)
    return {"ok": True, "task": task.model_dump()}


def _excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def _write_table(ws: Any, title: str, rows: list[Any]) -> None:
    ws.append([title])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor="1E3A8A")
    if not rows:
        ws.append(["（无数据）"])
        ws.append([])
        return
    if isinstance(rows[0], dict):
        headers = list(rows[0].keys())
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([_excel_value(row.get(header)) for header in headers])
    else:
        for row in rows:
            ws.append([_excel_value(row)])
    ws.append([])


def _write_mapping(ws: Any, title: str, data: dict[str, Any]) -> None:
    ws.append([title])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor="1E3A8A")
    for key, value in data.items():
        if isinstance(value, list):
            _write_table(ws, str(key), value)
        elif isinstance(value, dict):
            ws.append([key, json.dumps(value, ensure_ascii=False)])
        else:
            ws.append([key, _excel_value(value)])
    ws.append([])


def _paper_workbook_bytes(paper: ObjectInstance, engagement: ObjectInstance) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    paper_data = dict(paper.data or {})
    sheet_data = paper_data.get("sheet_data") or {}
    if not sheet_data:
        sheet_data = {"底稿": {"状态": "未生成"}}
    for raw_name, section in sheet_data.items():
        if raw_name in {"preparer", "prepared_at", "reviewer", "reviewed_at"}:
            continue
        title = re.sub(r"[:\\/?*\[\]]", "_", str(raw_name))[:31] or "Sheet"
        ws = wb.create_sheet(title)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.append([
            engagement.display_name,
            paper.display_name,
            (engagement.data or {}).get("period_end") or (engagement.data or {}).get("period"),
        ])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F172A")
        if isinstance(section, list):
            _write_table(ws, paper.display_name, section)
        elif isinstance(section, dict):
            rows = section.get("rows")
            if isinstance(rows, list):
                _write_table(ws, paper.display_name, rows)
            else:
                _write_mapping(ws, paper.display_name, section)
        else:
            ws.append([raw_name, _excel_value(section)])
        thin = Side(style="thin", color="CBD5E1")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in ws.columns:
            letter = column[0].column_letter
            width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 42)
            ws.column_dimensions[letter].width = width
    if not wb.sheetnames:
        wb.create_sheet("底稿")
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _paper_for_project(s: Session, engagement_code: str, paper_index: str) -> ObjectInstance:
    paper = next(
        (
            item
            for item in _objects_for_project(s, "WorkingPaper", engagement_code)
            if (item.data or {}).get("index") == paper_index
        ),
        None,
    )
    if not paper:
        raise HTTPException(404, f"底稿 {paper_index} 不存在")
    return paper


def _download_response(content: bytes, filename: str, media_type: str) -> StreamingResponse:
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/projects/{engagement_code}/papers/{paper_index}/export")
def export_paper(
    engagement_code: str,
    paper_index: str,
    s: Session = Depends(get_session),
) -> StreamingResponse:
    engagement = _require_engagement(s, engagement_code)
    paper = _paper_for_project(s, engagement_code, paper_index)
    filename = f"{paper_index}_{(paper.data or {}).get('name') or paper.display_name}.xlsx"
    return _download_response(
        _paper_workbook_bytes(paper, engagement),
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/projects/{engagement_code}/export")
def export_project(engagement_code: str, s: Session = Depends(get_session)) -> StreamingResponse:
    engagement = _require_engagement(s, engagement_code)
    papers = _objects_for_project(s, "WorkingPaper", engagement_code)
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        manifest = {
            "project": engagement.display_name,
            "engagement_code": engagement_code,
            "exported_at": _now(),
            "paper_count": len(papers),
        }
        archive.writestr("项目清单.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        for paper in papers:
            index = str((paper.data or {}).get("index") or paper.id)
            name = _safe_filename(str((paper.data or {}).get("name") or paper.display_name))
            archive.writestr(f"底稿/{index}_{name}.xlsx", _paper_workbook_bytes(paper, engagement))
    filename = f"{(engagement.data or {}).get('client_name') or engagement.display_name}_年审底稿包.zip"
    return _download_response(buffer.getvalue(), filename, "application/zip")
