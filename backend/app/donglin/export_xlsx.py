"""把已填底稿的数据写回真实东林 WP_FSR.xlsm 模板副本，返回 xlsx 文件给用户下载。

工作流:
  1. 拷贝 WP_FSR_donglin.xlsm → 临时文件
  2. 按 cell_mapping 把 sheet_data 的值写入对应 sheet!cell
  3. 返回拷贝后的 xlsx 二进制
"""
from __future__ import annotations
import json
import shutil
import tempfile
from pathlib import Path

from sqlmodel import Session, select
from fastapi import HTTPException
from fastapi.responses import FileResponse

ROOT = Path(__file__).resolve().parent.parent.parent
TEMPLATE_FILE = ROOT / "data" / "donglin" / "templates" / "WP_FSR_donglin.xlsm"
LAYOUT_FILE = ROOT / "data" / "donglin" / "templates" / "template_layout.json"


# 简化版 cell_mapping (服务端独立维护一份)
SUMMARY_CELL_MAP = {
    # A1 货币资金（按真实模板 A1 sheet 行 9-13 写入 D 列 = "期末未审金额"）
    "A1": {
        "tb_balance":           ("A1", "G9"),   # 审核确认额（明细账）
        "book_balance_total":   ("A1", "D9"),   # 期末未审金额
        "tb_diff":              ("A1", "H9"),   # 备注
    },
    # A6 应收账款
    "A6": {
        "tb_closing_unaudited": ("A6", "D9"),
        "aux_dr_total":         ("A6", "E9"),
        "aux_cr_total":         ("A6", "F9"),
        "closing_audited":      ("A6", "G9"),
        "reclass_to_advance":   ("A6", "H11"),  # 备注
    },
    "A9": {
        "tb_balance":           ("A9", "D9"),
        "reclass_to_tax":       ("A9", "H11"),
    },
    "A24": {
        "opening_cost":         ("A24", "C9"),
        "closing_cost":         ("A24", "D9"),
        "book_depreciation":    ("A24", "C11"),
        "recomputed_depreciation": ("A24", "D11"),
        "depreciation_diff":    ("A24", "E11"),
    },
    "B1": {
        "tb_balance":           ("B1", "D9"),
        "weighted_avg_rate":    ("B1", "H9"),
    },
}

# customer/asset/loan detail 的列映射（写入对应明细 sheet 的 B~J 列，行 8 起）
DETAIL_COL_MAP = {
    # paper_code → (detail_sheet_name, {field: col_letter})
    "A6_customer_detail": ("A6-2", {
        "customer_code": "B",
        "customer_name": "C",
        "closing_dr":    "G",   # 真实 A6-2 第 G 列 = 借方累计
        "closing_cr":    "H",   # 第 H 列 = 贷方累计
        "classification": "J",
    }),
    "A1_bank_detail": ("A1-2", {
        "account_no":     "B",
        "bank_name":      "B",
        "book_balance":   "F",
        "confirmation_balance": "I",
    }),
    "A9_customer_detail": ("A9-2", {
        "customer_code": "B",
        "customer_name": "C",
        "closing_dr":    "G",
        "closing_cr":    "H",
    }),
    "A24_asset_detail": ("A24-2", {
        "asset_class":  "B",
        "avg_cost":     "C",
        "years":        "D",
        "recomputed":   "E",
    }),
    "B1_loan_detail": ("B1-2", {
        "creditor":    "B",
        "principal":   "C",
        "rate":        "D",
        "term_start":  "E",
        "term_end":    "F",
        "guarantor":   "G",
    }),
}


def export_paper_xlsx(paper_code: str, session: Session) -> Path:
    """把 paper_code 对应底稿数据写入 .xlsm 副本，返回临时文件路径。"""
    if not TEMPLATE_FILE.exists():
        raise HTTPException(500, f"模板文件不存在: {TEMPLATE_FILE}")

    import openpyxl
    from ..models import ObjectInstance

    # 1. 查询 paper 的 sheet_data
    wp = next((o for o in session.exec(select(ObjectInstance).where(
        ObjectInstance.type_code == "WorkingPaper"))
        if (o.data or {}).get("index") == paper_code), None)
    if not wp:
        raise HTTPException(404, f"底稿 {paper_code} 未找到")

    sheet_data = (wp.data or {}).get("sheet_data") or {}
    if not sheet_data:
        raise HTTPException(400, f"底稿 {paper_code} 尚未填稿")

    # 2. 拷贝模板到临时文件
    tmp_dir = Path(tempfile.gettempdir())
    out_path = tmp_dir / f"东林审计_江苏大王_{paper_code}_{wp.id}.xlsm"
    shutil.copy(str(TEMPLATE_FILE), str(out_path))

    # 3. 打开拷贝并写入
    wb = openpyxl.load_workbook(str(out_path), keep_vba=True)

    # 3a. summary 字段写到主表 (A1/A6/A9/A24/B1)
    summary_map = SUMMARY_CELL_MAP.get(paper_code, {})
    summary_data = sheet_data.get("summary", {})
    written = 0
    for field, value in summary_data.items():
        if not isinstance(value, (int, float, str)):
            continue
        if field in summary_map:
            sheet_name, cell = summary_map[field]
            if sheet_name in wb.sheetnames:
                wb[sheet_name][cell] = value
                written += 1

    # 3b. customer/asset/loan detail 写到明细表
    for sheet_key, rows_data in sheet_data.items():
        if not isinstance(rows_data, dict):
            continue
        rows = rows_data.get("rows", [])
        if not isinstance(rows, list) or not rows:
            continue
        # 查找对应的明细映射
        map_key = f"{paper_code}_{sheet_key}"
        if map_key not in DETAIL_COL_MAP:
            continue
        detail_sheet_name, col_map = DETAIL_COL_MAP[map_key]
        if detail_sheet_name not in wb.sheetnames:
            continue
        ws = wb[detail_sheet_name]
        # 真实模板从第 8 行起是数据
        start_row = 8
        max_rows = min(len(rows), 500)  # 防止过长
        for i, row in enumerate(rows[:max_rows]):
            excel_row = start_row + i
            for field, col_letter in col_map.items():
                if field in row and row[field] is not None:
                    ws[f"{col_letter}{excel_row}"] = row[field]
                    written += 1

    # 3c. 填头部元数据（编制人、复核人、日期）
    d = wp.data or {}
    preparer = d.get("filled_by", "王叙超")
    filled_at = (d.get("filled_at") or "2026-02-28")[:10]
    for sheet_name in [paper_code, f"{paper_code}-2"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                ws["H1"] = preparer
                ws["H2"] = filled_at
                ws["H3"] = "侯佳成"
                ws["H4"] = "2026-03-02"
            except Exception:
                pass

    wb.save(str(out_path))
    wb.close()
    return out_path
