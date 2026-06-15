"""从真实东林 WP_FSR.xlsm 抽取 5 张 demo 底稿的实际表头/起始行/列结构。

输出: backend/data/donglin/templates/template_layout.json

格式:
{
  "A6": {
    "sheet_name": "A6",
    "title": "应收账款",
    "meta_rows": [1, 2, 3, 4, 5],     # 元数据 5 行
    "title_row": 6,                    # sheet 大标题行
    "header_row": 7,                   # 真实表头行
    "data_start_row": 8,               # 数据起始
    "columns": [                       # 真实列结构
      {"letter": "A", "label": "索引号"},
      {"letter": "B", "label": "项目"},
      ...
    ]
  },
  ...
}
"""
from __future__ import annotations
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent  # backend/
TEMPLATE_FILE = ROOT / "data" / "donglin" / "templates" / "WP_FSR_donglin.xlsm"
OUT_FILE = ROOT / "data" / "donglin" / "templates" / "template_layout.json"

# 5 张 demo 底稿 + 它们的子表
DEMO_SHEETS = [
    # paper_code, sheet_name, sub_label, kind
    ("A1",  "A1",   "审定表",        "summary"),
    ("A1",  "A1-2", "明细表",        "detail"),
    ("A6",  "A6",   "审定表",        "summary"),
    ("A6",  "A6-2", "客户明细",      "detail"),
    ("A6",  "A6-3", "函证明细",      "detail"),
    ("A9",  "A9",   "审定表",        "summary"),
    ("A9",  "A9-2", "明细表",        "detail"),
    ("A24", "A24",  "审定表",        "summary"),
    ("A24", "A24-2","明细表",        "detail"),
    ("B1",  "B1",   "审定表",        "summary"),
    ("B1",  "B1-2", "借款明细",      "detail"),
]


def extract_one_sheet(ws) -> dict:
    """对单个 sheet 抽取布局。"""
    out = {"sheet_name": ws.title}

    # 1) 收集前 10 行的非空单元格
    rows_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=12)):
        non_empty = []
        for j, v in enumerate(row[:15]):
            if v is not None:
                non_empty.append({"col": chr(65 + j), "value": str(v)[:60]})
        rows_data.append({"row": i + 1, "cells": non_empty})

    # 2) 找 header_row：通常是包含 "索引号" 或 "项目" 的行
    header_row_num = None
    header_cells = []
    for r in rows_data:
        labels = [c["value"] for c in r["cells"]]
        joined = " ".join(labels)
        if "索引号" in joined or ("项" in joined and "目" in joined and len(r["cells"]) >= 4):
            header_row_num = r["row"]
            header_cells = r["cells"]
            break

    if not header_row_num:
        # fallback: 取第 7 行
        header_row_num = 7
        header_cells = next((r["cells"] for r in rows_data if r["row"] == 7), [])

    # 3) 解析 title (通常在 header 前一行，居中)
    title = ""
    for r in rows_data:
        if r["row"] == header_row_num - 1 and r["cells"]:
            title = r["cells"][0]["value"]
            break

    # 4) 估算数据行数（采样后 100 行看哪一行后非空）
    data_start = header_row_num + 1
    data_count = 0
    for row in ws.iter_rows(min_row=data_start, max_row=data_start + 1000, values_only=True):
        if any(v is not None for v in row[:8]):
            data_count += 1
        else:
            if data_count > 0:
                break

    out.update({
        "title": title,
        "meta_rows": list(range(1, header_row_num - 1)),
        "title_row": header_row_num - 1,
        "header_row": header_row_num,
        "data_start_row": data_start,
        "data_row_count_estimate": data_count,
        "columns": [
            {"letter": c["col"], "label": c["value"]}
            for c in header_cells
        ],
        "raw_meta_preview": rows_data,
    })
    return out


def main():
    print(f"[extract_template] 打开 {TEMPLATE_FILE}")
    wb = openpyxl.load_workbook(str(TEMPLATE_FILE), data_only=True,
                                read_only=True, keep_vba=False)
    available = set(wb.sheetnames)

    layout = {}
    for paper_code, sheet_name, sub_label, kind in DEMO_SHEETS:
        if sheet_name not in available:
            print(f"  ⚠ {sheet_name} 不存在，跳过")
            continue
        ws = wb[sheet_name]
        info = extract_one_sheet(ws)
        info["paper_code"] = paper_code
        info["sub_label"] = sub_label
        info["kind"] = kind
        # 用 sheet_name 作 key
        layout[sheet_name] = info
        print(f"  ✓ {sheet_name} ({info['title']}) — {len(info['columns'])} 列 · "
              f"~{info['data_row_count_estimate']} 数据行")

    OUT_FILE.write_text(
        json.dumps(layout, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"[extract_template] ✓ 写入 {OUT_FILE}")
    return layout


if __name__ == "__main__":
    main()
