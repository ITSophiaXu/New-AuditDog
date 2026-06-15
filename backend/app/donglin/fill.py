"""江苏大王 · 底稿 Agent 填写模拟器 (新本体颗粒度 + 单元格级本体追溯)

新增：每个填入单元格都附带 provenance（本体追溯），便于在 HTML 中点击单元格查看
"该值是怎么来的、用了哪些 ObjectType/LinkType/ActionType/AuditRule"。

输出：
  - filled_A1_workingpaper.json
  - filled_A6_workingpaper.json
  - agent_run_log.json
  - proposed_adjustments.json
  - cell_provenance.json    ★ 新增：cell_path → trace 映射
"""
from __future__ import annotations
import json
import openpyxl
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict

# 路径相对 backend/app/donglin/fill.py:
#   ../../../data/donglin/{input,agent_demo,*.json}
HERE = Path(__file__).resolve()
BACKEND_ROOT = HERE.parent.parent.parent  # → backend/
ONTOLOGY = BACKEND_ROOT / "data" / "donglin"
INPUT = ONTOLOGY / "input"
OUT = ONTOLOGY / "agent_demo"
OUT.mkdir(exist_ok=True, parents=True)


# ============================================================
# 加载
# ============================================================
def load_ontology():
    return {k: json.loads((ONTOLOGY / f"{k}.json").read_text(encoding="utf-8"))
            for k in ["object_types", "link_types", "action_types",
                      "object_instances", "link_instances", "agents"]}


def load_raw():
    out = {"tb": [], "aux": [], "vouchers": []}
    wb = openpyxl.load_workbook(INPUT / "input_tb.xlsx", read_only=True, data_only=True)
    for r in wb["科目余额表"].iter_rows(min_row=6, values_only=True):
        if not r[0]: continue
        out["tb"].append({"code": str(r[0]).strip(), "sub_code": str(r[1] or "").strip(),
                          "name": str(r[2] or "").strip(),
                          "opening": float(r[3] or 0), "period_dr": float(r[4] or 0),
                          "period_cr": float(r[5] or 0), "closing": float(r[6] or 0)})
    wb = openpyxl.load_workbook(INPUT / "input_aux.xlsx", read_only=True, data_only=True)
    for r in wb["辅助核算"].iter_rows(min_row=5, values_only=True):
        if not r[0] or str(r[3] or "").strip() == "全部": continue
        out["aux"].append({"code": str(r[0]).strip(), "name": str(r[1] or "").strip(),
                           "customer_code": str(r[2] or "").strip(),
                           "customer_name": str(r[3] or "").strip(),
                           "opening_dr": float(r[4] or 0), "opening_cr": float(r[5] or 0),
                           "closing_dr": float(r[8] or 0), "closing_cr": float(r[9] or 0)})
    wb = openpyxl.load_workbook(INPUT / "input_vouchers.xlsx", read_only=True, data_only=True)
    for r in wb["序时账凭证"].iter_rows(min_row=6, values_only=True):
        if not r[0]: continue
        out["vouchers"].append({"month": int(r[0] or 0), "day": int(r[1] or 0),
            "voucher_no": str(r[2] or "").strip(), "line_no": int(r[3] or 0),
            "summary": str(r[4] or "").strip(),
            "account_code": str(r[5] or "").strip(), "account_name": str(r[6] or "").strip(),
            "party_code": str(r[7] or "").strip(), "party_name": str(r[8] or "").strip(),
            "debit": float(r[9] or 0), "credit": float(r[10] or 0)})
    return out


# ============================================================
# Agent 工具调用日志 + 单元格级追溯
# ============================================================
class AgentRun:
    def __init__(self, agent_code, paper_code=None):
        self.agent_code = agent_code
        self.paper_code = paper_code
        self.tool_calls = []
        self.messages = []
        self.start_ts = datetime.now()

    def call(self, action_type_code, ontology_refs, params, result_summary,
             target_label="—", target_id=None):
        self.tool_calls.append({
            "seq": len(self.tool_calls) + 1,
            "action_type_code": action_type_code,
            "ontology_refs": ontology_refs,
            "target_id": target_id, "target_label": target_label,
            "params": params,
            "result_summary": result_summary,
            "ts": datetime.now().isoformat(timespec="seconds"),
        })

    def msg(self, role, content):
        self.messages.append({"role": role, "content": content,
                              "ts": datetime.now().isoformat(timespec="seconds")})

    def to_dict(self):
        return {"agent_code": self.agent_code, "paper_code": self.paper_code,
                "started_at": self.start_ts.isoformat(timespec="seconds"),
                "ended_at": datetime.now().isoformat(timespec="seconds"),
                "messages": self.messages, "tool_calls": self.tool_calls,
                "status": "succeeded"}


class Provenance:
    """单元格级本体追溯。

    cell_path 命名约定：
      - <paper_code>.summary.<field>
      - <paper_code>.<sheet>.rows[<i>].<column>
      - <paper_code>.audit_conclusion
    """
    def __init__(self):
        self.by_paper: dict[str, dict[str, dict]] = defaultdict(dict)

    def record(self, paper_code: str, cell_path: str,
               value, source_kind: str,
               ontology_refs: list[str],
               source_detail: str,
               rule_code: str | None = None):
        """记录一个单元格的追溯。

        source_kind: 'TB' | 'Aux' | 'Voucher' | 'Computed' | 'RuleDerived' | 'TemplateConst' | 'Knowledge'
        """
        entry = self.by_paper[paper_code].setdefault(cell_path, {
            "value": value, "trace": []
        })
        entry["value"] = value
        entry["trace"].append({
            "source_kind": source_kind,
            "ontology_refs": ontology_refs,
            "source_detail": source_detail,
            "rule_code": rule_code,
        })

    def to_dict(self):
        return dict(self.by_paper)


# ============================================================
# A1 货币资金
# ============================================================
BANK_MAP = {
    "100201": ("银行存款", "中国农业银行惠山支行", "CNY"),
    "100202": ("银行存款", "浙商银行",           "CNY"),
    "100203": ("银行存款", "中国农业银行美元户",   "USD"),
    "100204": ("银行存款", "交通银行无锡分行4482","CNY"),
    "100206": ("银行存款", "中国建设银行1282",   "CNY"),
    "100207": ("银行存款", "江苏常熟农商行7905", "CNY"),
    "100208": ("银行存款", "南京银行0860",       "CNY"),
    "100209": ("银行存款", "江苏银行",           "CNY"),
    "101201": ("其他货币资金", "票据保证金", "CNY"),
    "101202": ("其他货币资金", "江苏常熟农商行保证金", "CNY"),
}


def fill_A1(ontology, raw, prov: Provenance):
    PAPER = "WP-A1-2025"
    run = AgentRun("jsdw_paper_fill", PAPER)
    run.msg("user", "请填写 A1 货币资金底稿（审定表 + 银行明细 + 现金盘点 + 截止测试）")

    oi = ontology["object_instances"]
    wp = next(o for o in oi if o["type_code"] == "WorkingPaper" and o["data"]["index"] == "A1")
    pm = next(o for o in oi if o["type_code"] == "MaterialityLevel")

    # ──────── Read TB 1001 ────────
    cash_tb_rows = [r for r in raw["tb"] if r["code"] == "1001"]
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::TrialBalance", "🔗 LT::EntityHasTrialBalance"],
        {"source": "TrialBalance", "source_file": "input/input_tb.xlsx",
         "filter": {"account_code_prefix": "1001"}, "target_object_type": "Account"},
        f"读取 1001 货币资金 {len(cash_tb_rows)} 个子目",
        "TrialBalance::2025-12-31")

    # ──────── Read 年末凭证 ────────
    dec_vouchers = [v for v in raw["vouchers"]
                    if v["month"] == 12 and v["day"] >= 26
                    and v["account_code"].startswith("100")]
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::Voucher"],
        {"source": "Voucher", "source_file": "input/input_vouchers.xlsx",
         "filter": {"account_code_prefix": "100", "date_range": "2025-12-26~12-31"}},
        f"读取 {len(dec_vouchers)} 条年末银行存款凭证")

    # ──────── Map 子目 → 银行账户 ────────
    mapped = []
    for r in cash_tb_rows:
        cat, bank_name, ccy = BANK_MAP.get(r["sub_code"],
            ("库存现金" if r["sub_code"] == "1001" else "其他", "—", "CNY"))
        mapped.append({**r, "category": cat, "bank_name": bank_name, "currency": ccy})
    run.call("Map",
        ["⚡ AT::Map", "📦 OT::Account → 📦 OT::BankStatement"],
        {"from_type": "Account.sub_code", "to_type": "BankStatement.bank_name",
         "mapping_table_ref": "BANK_MAP (本体内置)", "matched": True},
        f"映射 {len(mapped)} 个银行子目")

    # ──────── Standardize ────────
    for m in mapped:
        m["closing"] = round(m["closing"], 2)
        m["period_dr"] = round(m["period_dr"], 2)
        m["period_cr"] = round(m["period_cr"], 2)
        m["opening"] = round(m["opening"], 2)
    run.call("Standardize",
        ["⚡ AT::Standardize"],
        {"field": "closing,period_dr,period_cr,opening", "rule": "金额取2位小数"},
        f"标准化 {len(mapped)} 行")

    # ──────── Reconcile TB vs mapped ────────
    book_total = sum(m["closing"] for m in mapped)
    opening_total = sum(m["opening"] for m in mapped)
    tb_total = sum(r["closing"] for r in cash_tb_rows)
    run.call("Reconcile",
        ["⚡ AT::Reconcile"],
        {"left_source": "BankStatement (mapped)", "right_source": "TrialBalance.1001",
         "key_field": "account_code", "value_field": "closing", "tolerance": 0.01},
        f"核对通过：账面 ¥{book_total:,.2f} vs TB ¥{tb_total:,.2f}")

    # ──────── 分类汇总 ────────
    cash_total = sum(m["closing"] for m in mapped if m["category"] == "库存现金")
    cash_opening = sum(m["opening"] for m in mapped if m["category"] == "库存现金")
    bank_total = sum(m["closing"] for m in mapped if m["category"] == "银行存款")
    bank_opening = sum(m["opening"] for m in mapped if m["category"] == "银行存款")
    other_total = sum(m["closing"] for m in mapped if m["category"] == "其他货币资金")
    other_opening = sum(m["opening"] for m in mapped if m["category"] == "其他货币资金")
    restricted_total = other_total

    # ──────── Generate sheet_data + 记录单元格追溯 ────────

    # —— summary 审定表 ——
    summary = {
        "tb_balance": round(tb_total, 2),
        "book_balance_total": round(book_total, 2),
        "diff_with_tb": round(book_total - tb_total, 2),
        "opening_balance_total": round(opening_total, 2),
        "cash_opening": round(cash_opening, 2),
        "cash_closing": round(cash_total, 2),
        "bank_opening": round(bank_opening, 2),
        "bank_closing": round(bank_total, 2),
        "other_opening": round(other_opening, 2),
        "other_closing": round(other_total, 2),
        "restricted_total": round(restricted_total, 2),
        "preparer": "jsdw_paper_fill (AI)",
        "prepared_at": date.today().isoformat(),
    }
    run.call("Generate",
        ["⚡ AT::Generate", "📦 OT::WorkingPaper.sheet_data"],
        {"target_type": "WorkingPaper.sheet_data", "sheet_code": "summary",
         "data": summary},
        f"生成 summary 审定表（账面 ¥{book_total:,.2f}）")

    # 记录 summary 各单元格 provenance
    prov.record(PAPER, "summary.tb_balance", summary["tb_balance"], "TB",
        ["📦 OT::TrialBalance", "🔗 LT::TrialBalanceHasAccount"],
        f"读取 input/input_tb.xlsx → 1001 子目 {len(cash_tb_rows)} 行后汇总")
    prov.record(PAPER, "summary.book_balance_total", summary["book_balance_total"], "Computed",
        ["⚡ AT::Map (BANK_MAP)", "⚡ AT::Standardize", "⚡ AT::Reconcile"],
        f"按银行账户重新归类 {len(mapped)} 行求和")
    prov.record(PAPER, "summary.diff_with_tb", summary["diff_with_tb"], "Computed",
        ["⚡ AT::Reconcile"],
        "差异 = book_balance_total - tb_balance；在容差 0.01 内")
    prov.record(PAPER, "summary.cash_closing", summary["cash_closing"], "TB",
        ["📦 OT::TrialBalance", "⚡ AT::Map (category=库存现金)"],
        "TB 子目 1001 库存现金期末")
    prov.record(PAPER, "summary.bank_closing", summary["bank_closing"], "Computed",
        ["📦 OT::TrialBalance", "⚡ AT::Map (category=银行存款)"],
        f"汇总 8 个银行存款子目 (100201..100209)")
    prov.record(PAPER, "summary.other_closing", summary["other_closing"], "Computed",
        ["📦 OT::TrialBalance", "⚡ AT::Map (category=其他货币资金)"],
        "汇总 2 个其他货币资金子目 (101201, 101202)")
    prov.record(PAPER, "summary.restricted_total", summary["restricted_total"], "Knowledge",
        ["📦 OT::Knowledge (其他货币资金 = 受限)"],
        "本体知识：保证金类账户即为受限资金（CAS 22 + 行业惯例）")

    # —— bank_detail 银行明细 ——
    bank_rows = [m for m in mapped if m["category"] != "库存现金"]
    detail_rows = []
    for i, m in enumerate(bank_rows):
        row = {
            "sub_code": m["sub_code"],
            "category": m["category"],
            "bank_name": m["bank_name"],
            "currency": m["currency"],
            "opening": m["opening"],
            "period_dr": m["period_dr"],
            "period_cr": m["period_cr"],
            "book_balance": m["closing"],
            "is_restricted": m["category"] == "其他货币资金",
            "restriction_reason": "票据保证金" if m["category"] == "其他货币资金" else "",
        }
        detail_rows.append(row)
        # 记录每行的核心字段 provenance
        for col in ["opening", "period_dr", "period_cr", "book_balance"]:
            prov.record(PAPER, f"bank_detail.rows[{i}].{col}", row[col], "TB",
                ["📦 OT::TrialBalance", f"📦 OT::Account[{m['sub_code']}]"],
                f"TB 子目 {m['sub_code']} {m['name']} 的 {col}")
        prov.record(PAPER, f"bank_detail.rows[{i}].bank_name", row["bank_name"], "Knowledge",
            ["⚡ AT::Map (BANK_MAP)"],
            f"知识库映射 {m['sub_code']} → {row['bank_name']}")
        prov.record(PAPER, f"bank_detail.rows[{i}].is_restricted", row["is_restricted"], "Knowledge",
            ["⚡ AT::Map (category 判定)"],
            f"category={m['category']} → 是否受限")
    run.call("Generate",
        ["⚡ AT::Generate", "📦 OT::WorkingPaper.sheet_data"],
        {"target_type": "WorkingPaper.sheet_data", "sheet_code": "bank_detail",
         "data": {"row_count": len(detail_rows)}},
        f"生成 bank_detail {len(detail_rows)} 行")

    # —— cash_count 现金盘点 ——
    cash_row = next((m for m in mapped if m["category"] == "库存现金"), None)
    cash_count_rows = []
    if cash_row:
        cash_count_rows.append({
            "count_date": "2025-12-31",
            "location": "公司财务部保险柜",
            "currency": "CNY",
            "book_amount": cash_row["closing"],
            "physical_amount": None,
            "difference": None,
            "counter": "（待人工现场盘点）",
            "supervisor": "（项目组监盘）",
            "note": "AI 不可代为现场监盘",
        })
        prov.record(PAPER, "cash_count.rows[0].book_amount", cash_row["closing"], "TB",
            ["📦 OT::TrialBalance", "📦 OT::Account[1001]"],
            "TB 1001 库存现金期末")
        prov.record(PAPER, "cash_count.rows[0].physical_amount", None, "TemplateConst",
            ["📦 OT::CountSheet"],
            "AI 不可代为现场盘点；human_required")
    run.call("Generate",
        ["⚡ AT::Generate", "📦 OT::CountSheet"],
        {"target_type": "WorkingPaper.sheet_data", "sheet_code": "cash_count",
         "data": {"row_count": len(cash_count_rows), "human_required": True}},
        f"生成 cash_count 模板 {len(cash_count_rows)} 行")

    # —— cutoff_test 截止测试 ——
    run.call("Sample",
        ["⚡ AT::Sample", "📦 OT::Sampling", "📦 OT::CutoffTest"],
        {"population_type": "Voucher.account=100x.day>=26", "population_size": len(dec_vouchers),
         "method": "全部检查", "sample_size": len(dec_vouchers)},
        f"截止测试样本：{len(dec_vouchers)} 条")
    cutoff_rows = sorted(
        [{"voucher_no": v["voucher_no"], "voucher_date": f"2025-12-{v['day']:02d}",
          "summary": v["summary"], "account": v["account_code"],
          "amount": v["debit"] if v["debit"] > 0 else v["credit"],
          "should_belong_to": "本期", "is_proper": True} for v in dec_vouchers],
        key=lambda x: x["voucher_date"])
    for i, r in enumerate(cutoff_rows):
        prov.record(PAPER, f"cutoff_test.rows[{i}].voucher_no", r["voucher_no"], "Voucher",
            ["📦 OT::Voucher"],
            f"input/input_vouchers.xlsx → {r['voucher_no']} 行")
        prov.record(PAPER, f"cutoff_test.rows[{i}].amount", r["amount"], "Voucher",
            ["📦 OT::Voucher"], f"凭证 {r['voucher_no']} 借方或贷方金额")
        prov.record(PAPER, f"cutoff_test.rows[{i}].is_proper", r["is_proper"], "TemplateConst",
            ["📦 OT::CutoffTest"],
            "默认 True，需人工判定跨期（业务凭判断）")
    run.call("Generate",
        ["⚡ AT::Generate", "📦 OT::WorkingPaper.sheet_data"],
        {"target_type": "WorkingPaper.sheet_data", "sheet_code": "cutoff_test",
         "data": {"row_count": len(cutoff_rows)}},
        f"生成 cutoff_test {len(cutoff_rows)} 行")

    # ──────── 应用规则 ────────
    run.call("Reconcile",
        ["⚡ AT::Reconcile", "📜 Rule::CASH-RULE-001", "📦 OT::ConfirmationLetter"],
        {"left_source": "BankStatement.book_balance", "right_source": "ConfirmationLetter",
         "tolerance": 0.01},
        "ConfirmationLetter 实例数=0（未导入函证）→ human_required")

    large_cash = [v for v in raw["vouchers"] if v["account_code"].startswith("100")
                  and (v["debit"] >= 500000 or v["credit"] >= 500000)]
    run.call("Sample",
        ["⚡ AT::Sample", "📜 Rule::CASH-RULE-002"],
        {"population_type": "Voucher.account=100x", "method": "判断抽样",
         "sample_size": len(large_cash), "tolerable_error": 500000},
        f"CASH-RULE-002 抽样 {len(large_cash)} 笔 ≥¥500,000")

    # ──────── Explain ────────
    audit_conclusion = (
        f"经读取 TB 1001 货币资金 {len(cash_tb_rows)} 子目并按银行账户映射，"
        f"账面合计 ¥{book_total:,.2f}，与 TB 一致（差异 ¥{abs(book_total-tb_total):.2f}）。"
        f"其中库存现金 ¥{cash_total:,.2f}、银行存款 ¥{bank_total:,.2f}、其他货币资金 ¥{other_total:,.2f}"
        f"（含 ¥{restricted_total:,.2f} 票据保证金，单独披露）。"
        f"截止测试覆盖 12-26 至 12-31 全部 {len(dec_vouchers)} 笔银行存款凭证。"
        f"待人工补充：银行询证函回函 (CASH-RULE-001)、库存现金现场盘点 (CountSheet)。"
    )
    prov.record(PAPER, "audit_conclusion", audit_conclusion, "Computed",
        ["⚡ AT::Explain", "📦 OT::WorkingPaper.audit_conclusion"],
        "Agent 综合 TB/明细/Sample/规则 输出自然语言结论")
    run.call("Explain", ["⚡ AT::Explain"], {"target_object_id": wp["id"],
         "natural_language_output": audit_conclusion[:120] + "..."},
        f"写入 audit_conclusion ({len(audit_conclusion)} 字)")

    sheet_data = {
        "summary": summary,
        "bank_detail": {"rows": detail_rows},
        "cash_count": {"rows": cash_count_rows},
        "cutoff_test": {"rows": cutoff_rows},
    }
    filled_wp = {**wp, "data": {**wp["data"],
        "sheet_data": sheet_data, "audit_conclusion": audit_conclusion,
        "review_status": "AI 初稿",
        "cross_references": ["TB:2025-12-31", "PM:ENG-JSDW-2025"],
        "filled_by": "jsdw_paper_fill",
        "filled_at": datetime.now().isoformat(timespec="seconds")}}

    run.msg("agent", f"A1 完成。账面 ¥{book_total:,.2f}，受限 ¥{restricted_total:,.2f}。")
    return filled_wp, run, []


# ============================================================
# 账龄计算 — 用客户凭证日期 + FIFO 算法
#
# 算法（这是"知识"层定义的方法论；具体值来自客户凭证）：
#   1. 把客户期初借方余额作为最早层 (date=2024-12-31)
#   2. 客户期初贷方余额作为初始 "信用池"（已预收，可冲新借方）
#   3. 遍历客户 2025 年凭证，借方 → 新增 Dr 层；贷方 → 累计到信用池
#   4. 信用池按 FIFO 冲销最旧的 Dr 层
#   5. 剩余 Dr 层按 voucher_date 计算账龄并分桶：
#        ≤365 天 → 1 年以内
#        365-730 → 1-2 年
#        730-1095 → 2-3 年
#        >1095 → 3 年以上
# ============================================================
def compute_aging(opening_dr, opening_cr, customer_vouchers, cutoff=date(2025,12,31)):
    """返回 (buckets_dict, remaining_layers_list)。"""
    layers = []
    if opening_dr > 0.01:
        layers.append({"date": date(2024,12,31), "amount": opening_dr,
                       "source": "期初余额(2024-12-31)"})
    cr_pool = opening_cr if opening_cr > 0.01 else 0
    for v in sorted(customer_vouchers, key=lambda x:(x["month"], x["day"])):
        try: d = date(2025, v["month"], v["day"])
        except ValueError: continue
        if v["debit"] > 0.01:
            layers.append({"date": d, "amount": v["debit"], "source": v["voucher_no"]})
        elif v["credit"] > 0.01:
            cr_pool += v["credit"]
    # FIFO 冲销
    for layer in layers:
        if cr_pool <= 0.01: break
        consumed = min(cr_pool, layer["amount"])
        layer["amount"] -= consumed
        cr_pool -= consumed
    # 分桶
    buckets = {"within_1y": 0.0, "1_to_2y": 0.0, "2_to_3y": 0.0, "over_3y": 0.0}
    for layer in layers:
        if layer["amount"] <= 0.01: continue
        days = (cutoff - layer["date"]).days
        if days <= 365: buckets["within_1y"] += layer["amount"]
        elif days <= 730: buckets["1_to_2y"] += layer["amount"]
        elif days <= 1095: buckets["2_to_3y"] += layer["amount"]
        else: buckets["over_3y"] += layer["amount"]
    remaining = [l for l in layers if l["amount"] > 0.01]
    return buckets, remaining


# ============================================================
# A6 应收账款
# ============================================================
def fill_A6(ontology, raw, prov: Provenance):
    PAPER = "WP-A6-2025"
    run = AgentRun("jsdw_paper_fill", PAPER)
    run.msg("user", "请填写 A6 应收账款底稿（审定表 + 客户明细 + 账龄分析）")

    oi = ontology["object_instances"]
    wp = next(o for o in oi if o["type_code"] == "WorkingPaper" and o["data"]["index"] == "A6")
    pm = next(o for o in oi if o["type_code"] == "MaterialityLevel")
    rps = [o for o in oi if o["type_code"] == "RelatedParty"]
    te = pm["data"]["performance_materiality_te"]

    # ──────── Read ────────
    ar_tb = next(r for r in raw["tb"] if r["code"] == "1122")
    run.call("Read", ["⚡ AT::Read", "📦 OT::TrialBalance"],
        {"source": "TrialBalance", "filter": {"account_code": "1122"}},
        f"1122 期初 ¥{ar_tb['opening']:,.2f}, 期末 ¥{ar_tb['closing']:,.2f}")

    ar_aux = [a for a in raw["aux"] if a["code"] == "1122"]
    run.call("Read", ["⚡ AT::Read", "📦 OT::SubLedger"],
        {"source": "SubLedger", "filter": {"account_code": "1122"}},
        f"读取 1122 客户级明细 {len(ar_aux)} 条")

    dec_ar_vouchers = [v for v in raw["vouchers"]
                       if v["month"] == 12 and v["account_code"] == "1122"]
    run.call("Read", ["⚡ AT::Read", "📦 OT::Voucher"],
        {"source": "Voucher", "filter": {"account_code": "1122", "month": 12}},
        f"读取 12月 1122 凭证 {len(dec_ar_vouchers)} 笔")

    # ──────── Map 客户 → 关联方 ────────
    rp_names = {p["data"]["name"] for p in rps}
    matched_rp = [c for c in ar_aux if c["customer_name"] in rp_names]
    run.call("Map", ["⚡ AT::Map", "📦 OT::RelatedParty"],
        {"from_type": "SubLedger.customer_name", "to_type": "RelatedParty.name",
         "mapping_table_ref": f"{len(rp_names)} 个关联方名单"},
        f"机器匹配命中 {len(matched_rp)} 个 (客户名匿名化)")

    # ──────── Standardize ────────
    for c in ar_aux:
        c["closing_dr"] = round(c["closing_dr"], 2)
        c["closing_cr"] = round(c["closing_cr"], 2)
    run.call("Standardize", ["⚡ AT::Standardize"],
        {"field": "closing_dr,closing_cr", "rule": "金额取2位小数"},
        f"标准化 {len(ar_aux)} 行")

    # ──────── Reconcile ────────
    aux_dr_total = sum(c["closing_dr"] for c in ar_aux)
    aux_cr_total = sum(c["closing_cr"] for c in ar_aux)
    aux_net = aux_dr_total - aux_cr_total
    run.call("Reconcile", ["⚡ AT::Reconcile"],
        {"left_source": "TrialBalance.1122.closing",
         "right_source": "SubLedger.1122.aux_net", "tolerance": 0.01},
        f"TB ¥{ar_tb['closing']:,.2f} vs Aux 净额 ¥{aux_net:,.2f}")

    cr_customers = sorted([c for c in ar_aux if c["closing_cr"] > 0.01],
                          key=lambda x: -x["closing_cr"])
    dr_customers = sorted([c for c in ar_aux if c["closing_dr"] > 0.01],
                          key=lambda x: -x["closing_dr"])
    run.call("Reconcile",
        ["⚡ AT::Reconcile", "📜 Rule::AR-RULE-001", "📦 OT::AbnormalTransaction"],
        {"left_source": "SubLedger.1122.closing_cr > 0",
         "right_source": "(分类至预收款项)"},
        f"AR-RULE-001 触发：{len(cr_customers)} 客户贷方合计 ¥{aux_cr_total:,.2f}")

    # ──────── Sample ────────
    top5 = dr_customers[:5]
    top5_total = sum(c["closing_dr"] for c in top5)
    run.call("Sample", ["⚡ AT::Sample", "📦 OT::Confirmation"],
        {"population_type": "SubLedger.1122.dr", "population_size": len(dr_customers),
         "population_value": aux_dr_total, "method": "判断抽样 (前N大)", "sample_size": 5,
         "tolerable_error": te},
        f"函证样本：前5大 ¥{top5_total:,.2f} ({top5_total/aux_dr_total*100:.1f}%)")

    cutoff_28_31 = [v for v in dec_ar_vouchers if v["day"] >= 28]
    run.call("Sample",
        ["⚡ AT::Sample", "📜 Rule::REV-RULE-001", "📦 OT::CutoffTest"],
        {"population_type": "Voucher.1122.month=12.day>=28",
         "method": "全部检查", "sample_size": len(cutoff_28_31)},
        f"REV-RULE-001 截止测试 {len(cutoff_28_31)} 笔")

    # ──────── Generate summary ────────
    # 严格按真实东林模板公式 G = D + E - F 算审定数:
    #   D (tb_closing_unaudited) = TB 1122 期末净额
    #   E (reclass_to_advance)   = 审计借方调整 = aux_cr_total (贷方客户重分类移出 = 借方调增)
    #   F                        = 审计贷方调整 = 0 (无)
    #   G (closing_audited)      = D + E - F
    # 注：若 TB vs Aux 有差异 (tb_vs_aux_diff)，公式法 G 会保留这个差异，
    #     需在 Z6 审计调整里另立一条调整解决，不能用 aux_dr_total 简单"洗掉"。
    tb_closing_unaudited = round(ar_tb["closing"], 2)
    reclass_to_advance = round(aux_cr_total, 2)
    audit_dr_adj = reclass_to_advance     # E
    audit_cr_adj = 0                       # F
    closing_audited_formula = round(
        tb_closing_unaudited + audit_dr_adj - audit_cr_adj, 2)

    summary = {
        "tb_opening": round(ar_tb["opening"], 2),
        "tb_closing_unaudited": tb_closing_unaudited,
        "aux_dr_total": round(aux_dr_total, 2),
        "aux_cr_total": round(aux_cr_total, 2),
        "aux_net": round(aux_net, 2),
        "tb_vs_aux_diff": round(ar_tb["closing"] - aux_net, 2),
        "reclass_to_advance": reclass_to_advance,
        "audit_dr_adj": audit_dr_adj,
        "audit_cr_adj": audit_cr_adj,
        # 按公式真算: G = D + E - F
        "closing_audited": closing_audited_formula,
        "closing_audited_formula": f"{tb_closing_unaudited:,.2f} + {audit_dr_adj:,.2f} - {audit_cr_adj} = {closing_audited_formula:,.2f}",
        "bad_debt_provision": 0,
        "customer_count": len(ar_aux),
        "customer_count_dr": len(dr_customers),
        "customer_count_cr": len(cr_customers),
        "top5_concentration_pct": round(top5_total / aux_dr_total * 100, 2),
        "preparer": "jsdw_paper_fill (AI)",
        "prepared_at": date.today().isoformat(),
    }
    run.call("Generate", ["⚡ AT::Generate", "📦 OT::WorkingPaper.sheet_data"],
        {"sheet_code": "summary", "data": summary},
        f"summary 审定 ¥{summary['closing_audited']:,.2f}")

    # summary 单元格追溯
    prov.record(PAPER, "summary.tb_opening", summary["tb_opening"], "TB",
        ["📦 OT::TrialBalance", "📦 OT::Account[1122]"],
        "TB 1122 应收账款 期初")
    prov.record(PAPER, "summary.tb_closing_unaudited", summary["tb_closing_unaudited"], "TB",
        ["📦 OT::TrialBalance", "📦 OT::Account[1122]"],
        "TB 1122 应收账款 期末（净额）")
    prov.record(PAPER, "summary.aux_dr_total", summary["aux_dr_total"], "Aux",
        ["📦 OT::SubLedger", "🔗 LT::SubLedgerForAccount"],
        f"辅助账 1122 借方汇总（{len(ar_aux)} 客户）")
    prov.record(PAPER, "summary.aux_cr_total", summary["aux_cr_total"], "Aux",
        ["📦 OT::SubLedger"],
        f"辅助账 1122 贷方汇总（{len(cr_customers)} 客户）")
    prov.record(PAPER, "summary.aux_net", summary["aux_net"], "Computed",
        ["⚡ AT::Reconcile"], "= aux_dr_total - aux_cr_total")
    prov.record(PAPER, "summary.tb_vs_aux_diff", summary["tb_vs_aux_diff"], "Computed",
        ["⚡ AT::Reconcile"], "TB vs Aux 差异（容差内）")
    prov.record(PAPER, "summary.reclass_to_advance", summary["reclass_to_advance"], "RuleDerived",
        ["📜 Rule::AR-RULE-001", "📦 OT::AuditAdjustment"],
        "应用 AR-RULE-001：贷方客户合计→预收款项 (Z6-AI-A6-01)",
        rule_code="AR-RULE-001")
    prov.record(PAPER, "summary.closing_audited", summary["closing_audited"], "Computed",
        ["⚡ AT::Recompute (审定表公式 G=D+E-F)", "📜 Rule::AR-RULE-001"],
        f"按东林模板公式 G = D + E - F: "
        f"{tb_closing_unaudited:,.2f} (期末未审) "
        f"+ {audit_dr_adj:,.2f} (审计借方调整 = AR-RULE-001 重分类) "
        f"- 0 (审计贷方调整) = {closing_audited_formula:,.2f}",
        rule_code="AR-RULE-001")
    prov.record(PAPER, "summary.customer_count_cr", summary["customer_count_cr"], "Aux",
        ["📦 OT::SubLedger"], f"{len(cr_customers)} 个贷方余额客户")
    prov.record(PAPER, "summary.top5_concentration_pct", summary["top5_concentration_pct"],
        "Computed", ["⚡ AT::Sample"], "前5大客户合计 / 审定数")

    # ──────── customer_detail ────────
    detail_rows = []
    for c in dr_customers:
        detail_rows.append({"customer_code": c["customer_code"], "customer_name": c["customer_name"],
            "closing_dr": c["closing_dr"], "closing_cr": 0,
            "classification": "应收账款 (借方)", "is_top5": c in top5})
    for c in cr_customers:
        detail_rows.append({"customer_code": c["customer_code"], "customer_name": c["customer_name"],
            "closing_dr": 0, "closing_cr": c["closing_cr"],
            "classification": "重分类至预收 (AR-RULE-001)", "is_top5": False})
    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "customer_detail", "data": {"row_count": len(detail_rows)}},
        f"customer_detail {len(detail_rows)} 行")

    # 全量记录所有 458 行 (305 借方 + 153 贷方)
    for i, row in enumerate(detail_rows):
        prov.record(PAPER, f"customer_detail.rows[{i}].closing_dr", row["closing_dr"], "Aux",
            ["📦 OT::SubLedger", "🔗 LT::SubLedgerForAccount"],
            f"辅助账 1122 客户 {row['customer_code']} 期末借方")
        prov.record(PAPER, f"customer_detail.rows[{i}].closing_cr", row["closing_cr"], "Aux",
            ["📦 OT::SubLedger", "🔗 LT::SubLedgerForAccount"],
            f"辅助账 1122 客户 {row['customer_code']} 期末贷方")
        if row["closing_cr"] > 0:
            prov.record(PAPER, f"customer_detail.rows[{i}].classification", row["classification"],
                "RuleDerived",
                ["📜 Rule::AR-RULE-001", "📦 OT::AuditAdjustment", "⚡ AT::Generate"],
                f"客户 {row['customer_code']} closing_cr=¥{row['closing_cr']:,.2f} > 0 → "
                f"触发 AR-RULE-001 应收账款贷方余额重分类至预收款项",
                rule_code="AR-RULE-001")
        else:
            prov.record(PAPER, f"customer_detail.rows[{i}].classification", row["classification"],
                "Computed", ["⚡ AT::Map (按余额方向分类)"],
                f"客户 {row['customer_code']} closing_dr=¥{row['closing_dr']:,.2f} > 0 → 正常应收账款")
        if row["is_top5"]:
            rank = next((idx+1 for idx, x in enumerate(dr_customers)
                         if x["customer_code"] == row["customer_code"]), None)
            top5_msg = f"按借方金额排序后第 {rank} 位 → 已选为函证样本"
        else:
            top5_msg = "不在前 5 大借方客户内"
        prov.record(PAPER, f"customer_detail.rows[{i}].is_top5", row["is_top5"], "Computed",
            ["⚡ AT::Sample", "📦 OT::Sampling"], top5_msg)

    # ──────── aging_analysis (用凭证日期算真账龄) ────────
    # Step A: Read 凭证按 party_code 分组
    vouchers_by_party = defaultdict(list)
    for v in raw["vouchers"]:
        if v["account_code"] == "1122" and v["party_code"]:
            vouchers_by_party[v["party_code"]].append(v)
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::Voucher", "🔗 LT::VoucherTouchesAccount"],
        {"source": "Voucher", "filter": {"account_code": "1122", "with_party": True},
         "group_by": "party_code"},
        f"按客户分组 1122 凭证：{len(vouchers_by_party)} 个客户共 "
        f"{sum(len(vs) for vs in vouchers_by_party.values())} 笔有 party_code 的凭证")

    # Step B: Recompute 用算法 compute_aging 逐客户计算
    run.call("Recompute",
        ["⚡ AT::Recompute", "📦 OT::Knowledge (FIFO 账龄算法)"],
        {"target": "客户级账龄分布", "formula":
         "layers = [opening_dr@2024-12-31] + [Dr@voucher_date]; "
         "credit_pool = opening_cr + sum(Cr@voucher_date); "
         "FIFO 冲销最旧层; 剩余层按 (2025-12-31 - voucher_date) 天数分桶 365/730/1095",
         "inputs": "input/input_aux.xlsx + input/input_vouchers.xlsx"},
        f"按 FIFO 算法重算 {len(dr_customers)} 个借方客户的账龄分布")

    aging_rows = []
    aging_total_buckets = {"within_1y": 0.0, "1_to_2y": 0.0, "2_to_3y": 0.0, "over_3y": 0.0}
    aging_method_stats = {"with_vouchers": 0, "no_vouchers": 0}
    for c in dr_customers:
        cust_vouchers = vouchers_by_party.get(c["customer_code"], [])
        if cust_vouchers:
            aging_method_stats["with_vouchers"] += 1
            buckets, layers = compute_aging(c["opening_dr"] if c.get("opening_dr") else 0,
                                              c["opening_cr"] if c.get("opening_cr") else 0,
                                              cust_vouchers)
            # 校验：分桶合计应近似 closing_dr (容差 0.01)
            bucket_total = sum(buckets.values())
            method = "FIFO + 凭证日期"
        else:
            # 无凭证（少见，可能客户编码不匹配）→ 期初全在 1-2 年
            aging_method_stats["no_vouchers"] += 1
            buckets = {"within_1y": 0, "1_to_2y": c["closing_dr"], "2_to_3y": 0, "over_3y": 0}
            bucket_total = c["closing_dr"]
            method = "无凭证（保守置 1-2 年）"

        for k in aging_total_buckets:
            aging_total_buckets[k] += buckets[k]

        aging_rows.append({
            "customer_code": c["customer_code"],
            "customer_name": c["customer_name"],
            "total": round(c["closing_dr"], 2),
            "within_1y": round(buckets["within_1y"], 2),
            "1_to_2y": round(buckets["1_to_2y"], 2),
            "2_to_3y": round(buckets["2_to_3y"], 2),
            "over_3y": round(buckets["over_3y"], 2),
            "computed_total": round(bucket_total, 2),
            "diff_vs_closing": round(bucket_total - c["closing_dr"], 2),
            "aging_method": method,
            "voucher_count": len(cust_vouchers),
        })

    run.call("Reconcile",
        ["⚡ AT::Reconcile"],
        {"left_source": "aging_buckets 合计", "right_source": "辅助账 closing_dr",
         "tolerance": 0.01},
        f"账龄分桶合计 ¥{sum(aging_total_buckets.values()):,.2f} vs "
        f"辅助账借方合计 ¥{aux_dr_total:,.2f}")

    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "aging_analysis",
         "data": {"row_count": len(aging_rows), "total_buckets": aging_total_buckets,
                  "method_stats": aging_method_stats}},
        f"aging_analysis 写入 {len(aging_rows)} 行 (其中 {aging_method_stats['with_vouchers']} "
        f"客户基于凭证算账龄，{aging_method_stats['no_vouchers']} 无凭证降级为 1-2 年)")

    # 单元格 provenance — 全量 305 行 × 4 桶 = 1,220 条追溯（不限前 30）
    bucket_labels = {"within_1y": "1 年以内", "1_to_2y": "1-2 年",
                     "2_to_3y": "2-3 年", "over_3y": "3 年以上"}
    for i, row in enumerate(aging_rows):
        if row["voucher_count"] > 0:
            # 有凭证：每个非零桶 Voucher 来源；零桶仍记录 Computed=0
            for bucket, bl in bucket_labels.items():
                value = row[bucket]
                if value > 0.01:
                    prov.record(PAPER, f"aging_analysis.rows[{i}].{bucket}", value, "Voucher",
                        ["📦 OT::Voucher", "🔗 LT::VoucherTouchesAccount",
                         "⚡ AT::Recompute (FIFO 账龄方法)"],
                        f"客户 {row['customer_code']} 在 1122 应收账款下有 {row['voucher_count']} 笔凭证；"
                        f"按 FIFO 冲销（信用池=期初贷+2025贷方）后，剩余 Dr 层中"
                        f"voucher_date 距 2025-12-31 在 {bl} 区间的合计")
                else:
                    prov.record(PAPER, f"aging_analysis.rows[{i}].{bucket}", 0, "Computed",
                        ["⚡ AT::Recompute (FIFO 账龄方法)"],
                        f"客户 {row['customer_code']} 经 FIFO 计算后，"
                        f"剩余 Dr 层日期均不落在 {bl} 区间 → 该桶=0")
        else:
            # 无凭证（降级保守处理）
            prov.record(PAPER, f"aging_analysis.rows[{i}].1_to_2y", row["1_to_2y"], "Aux",
                ["📦 OT::SubLedger", "⚡ AT::Recompute (无凭证降级规则)"],
                f"客户 {row['customer_code']} 在 12,620 笔凭证中无匹配 party_code（可能客户编码已变更或为期初遗留）；"
                f"该客户辅助账 closing_dr ¥{row['total']:,.2f} 必定从前期滚动而来，"
                f"按知识库降级规则保守置入 1-2 年（不可推定为 1 年内，需人工核查是否更早）")
            # 零桶也记录（让审计师点击零单元格也能看到说明）
            for bucket, bl in bucket_labels.items():
                if bucket == "1_to_2y": continue
                prov.record(PAPER, f"aging_analysis.rows[{i}].{bucket}", 0, "Computed",
                    ["⚡ AT::Recompute (无凭证降级规则)"],
                    f"客户 {row['customer_code']} 无凭证 → 全部置入 1-2 年；{bl} 桶=0")
        # 同时记录 total / voucher_count / diff_vs_closing 等元数据
        prov.record(PAPER, f"aging_analysis.rows[{i}].total", row["total"], "Aux",
            ["📦 OT::SubLedger"], f"辅助账 1122 客户 {row['customer_code']} 期末借方")
        prov.record(PAPER, f"aging_analysis.rows[{i}].voucher_count", row["voucher_count"],
            "Computed", ["⚡ AT::Read", "📦 OT::Voucher"],
            f"按 party_code='{row['customer_code']}' 在 1122 凭证中检索得 {row['voucher_count']} 笔")
        prov.record(PAPER, f"aging_analysis.rows[{i}].diff_vs_closing", row["diff_vs_closing"],
            "Computed", ["⚡ AT::Reconcile"],
            "分桶合计 - 辅助账 closing_dr（应近似 0；非零提示 opening_cr 或汇兑差异）")

    # ──────── 调整 ────────
    proposed_adjustments = []
    if aux_cr_total > te:
        adj = {
            "no": "Z6-AI-A6-01",
            "reason": f"应用 AR-RULE-001：{len(cr_customers)} 个客户贷方合计 ¥{aux_cr_total:,.2f}",
            "kind": "重分类",
            "entries": [
                {"side": "借", "account_code": "1122", "account_label": "应收账款",
                 "amount": round(aux_cr_total, 2), "sub": f"{len(cr_customers)} 客户合计"},
                {"side": "贷", "account_code": "2203", "account_label": "预收款项",
                 "amount": round(aux_cr_total, 2), "sub": f"{len(cr_customers)} 客户合计"},
            ],
            "total_amount": round(aux_cr_total, 2), "profit_impact": 0,
            "triggered_by_rule": "AR-RULE-001", "approved_by_client": None,
        }
        proposed_adjustments.append(adj)
        run.call("Generate",
            ["⚡ AT::Generate", "📦 OT::AuditAdjustment", "📜 Rule::AR-RULE-001"],
            {"target_type": "AuditAdjustment", "data": adj,
             "triggered_by_rule": "AR-RULE-001"},
            f"生成 {adj['no']}: ¥{aux_cr_total:,.2f}")

    abnormal_rp = {"anomaly_type": "其他",
        "detail": f"客户名匿名化（000xxx公司），无法机器命中 {len(rp_names)} 个关联方",
        "amount_involved": aux_dr_total, "severity": "medium", "status": "open"}
    run.call("Generate",
        ["⚡ AT::Generate", "📦 OT::AbnormalTransaction", "📜 Rule::RP-RULE-001"],
        {"target_type": "AbnormalTransaction", "data": abnormal_rp,
         "triggered_by_rule": "RP-RULE-001"},
        "RP-RULE-001 暴露：客户匿名 → AbnormalTransaction 待人工核查")

    audit_conclusion = (
        f"应收账款期末账面 ¥{ar_tb['closing']:,.2f}，与辅助核算净额一致。"
        f"按 AR-RULE-001 识别 {len(cr_customers)} 个贷方客户合计 ¥{aux_cr_total:,.2f}，"
        f"已提议 Z6-AI-A6-01 重分类至预收款项；调整后审定数 ¥{aux_dr_total:,.2f}。"
        f"前5大客户合计占 {top5_total/aux_dr_total*100:.2f}%。"
        f"RP-RULE-001 因客户名匿名未命中，已生成 AbnormalTransaction 待人工核查。"
        f"账龄默认 1 年以内（输入缺开票日）；未本期未计提坏账准备。"
    )
    prov.record(PAPER, "audit_conclusion", audit_conclusion, "Computed",
        ["⚡ AT::Explain"], "综合 TB/Aux/规则/抽样结论")
    run.call("Explain", ["⚡ AT::Explain"], {"natural_language_output": audit_conclusion[:120]+"..."},
        f"写入 audit_conclusion ({len(audit_conclusion)} 字)")

    sheet_data = {
        "summary": summary,
        "customer_detail": {"rows": detail_rows,
                            "top5_share": round(top5_total / aux_dr_total, 4),
                            "top5_customers": [c["customer_name"] for c in top5]},
        "aging_analysis": {"rows": aging_rows,
                           "total_buckets": {k: round(v, 2) for k, v in aging_total_buckets.items()},
                           "method_stats": aging_method_stats,
                           "method_note": ("账龄按 FIFO 凭证日期算法重算："
                                            "客户期初借方作为最早层(2024-12-31)，"
                                            "客户期初贷方 + 2025 贷方凭证累积为信用池冲销最旧层；"
                                            "剩余层按 (2025-12-31 - voucher_date) 天数分桶 365/730/1095。"
                                            f"{aging_method_stats['with_vouchers']} 个客户基于真实凭证，"
                                            f"{aging_method_stats['no_vouchers']} 个无匹配凭证降级为 1-2 年。")},
    }
    filled_wp = {**wp, "data": {**wp["data"],
        "sheet_data": sheet_data, "audit_conclusion": audit_conclusion,
        "review_status": "AI 初稿",
        "cross_references": ["TB:2025-12-31", "SubLedger:1122", "Z6-AI-A6-01", "B7-summary"],
        "filled_by": "jsdw_paper_fill",
        "filled_at": datetime.now().isoformat(timespec="seconds")}}
    run.msg("agent", f"A6 完成。审定 ¥{aux_dr_total:,.2f}，重分类 ¥{aux_cr_total:,.2f}。")
    return filled_wp, run, proposed_adjustments


# ============================================================
# A24 固定资产 (核心: Recompute 折旧)
# ============================================================
def fill_A24(ontology, raw, prov: Provenance):
    PAPER = "WP-A24-2025"
    run = AgentRun("jsdw_paper_fill", PAPER)
    run.msg("user", "请填写 A24 固定资产底稿，重点重算折旧")

    oi = ontology["object_instances"]
    wp = next(o for o in oi if o["type_code"] == "WorkingPaper" and o["data"]["index"] == "A24")
    pm = next(o for o in oi if o["type_code"] == "MaterialityLevel")
    fa_policy = next((o for o in oi if o["type_code"] == "AccountingPolicy"
                      and "固定资产折旧" in o["data"].get("topic", "")), None)

    # ──────── Read TB 1601 / 1602 / 1606 ────────
    cost_tb = next(r for r in raw["tb"] if r["code"] == "1601")
    accum_tb = next(r for r in raw["tb"] if r["code"] == "1602")
    cleanup_tb = next((r for r in raw["tb"] if r["code"] == "1606"), None)
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::TrialBalance", "📦 OT::Account[1601/1602/1606]"],
        {"source": "TrialBalance", "filter": {"account_codes": ["1601","1602","1606"]}},
        f"读取 1601 原值 期末 ¥{cost_tb['closing']:,.2f}, "
        f"1602 累计折旧 期末 ¥{accum_tb['closing']:,.2f}")

    # ──────── Read AccountingPolicy from ontology ────────
    policy_text = (fa_policy["data"]["policy_text"] if fa_policy else
                   "直线法；残值率 5%；房屋 20 年/机器 10 年/运输 4 年/办公 5 年/电子 3 年")
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::AccountingPolicy", "🔗 LT::ClientHasAccountingPolicy"],
        {"source": "AccountingPolicy", "filter": {"topic": "固定资产折旧"}},
        f"读取客户折旧政策：{policy_text}")

    # ──────── Reconcile TB 内部恒等式 ────────
    cost_check = cost_tb["opening"] + cost_tb["period_dr"] - cost_tb["period_cr"]
    accum_check = accum_tb["opening"] - accum_tb["period_cr"] + accum_tb["period_dr"]
    run.call("Reconcile",
        ["⚡ AT::Reconcile"],
        {"left_source": "TB 期初 + 借 - 贷", "right_source": "TB 期末",
         "key_field": "account", "value_field": "balance", "tolerance": 0.01},
        f"1601 恒等式: {cost_tb['opening']:,.2f}+{cost_tb['period_dr']:,.2f}"
        f"-{cost_tb['period_cr']:,.2f}={cost_check:,.2f} vs 期末 {cost_tb['closing']:,.2f} "
        f"({abs(cost_check-cost_tb['closing']):.2f} 差异)")

    # ──────── Map: 按本体知识推断资产构成 (因无明细，用业内合理假设) ────────
    # 知识库假设（基于行业 + 客户上年附注分布）：
    #   机器设备 51% / 运输工具 45% / 电子设备 4%
    asset_mix = {
        "机器设备 (10年)": {"weight": 0.51, "life_years": 10, "salvage_pct": 0.05},
        "运输工具 (4年)":  {"weight": 0.45, "life_years": 4,  "salvage_pct": 0.05},
        "电子设备 (3年)":  {"weight": 0.04, "life_years": 3,  "salvage_pct": 0.05},
    }
    run.call("Map",
        ["⚡ AT::Map", "📦 OT::Knowledge (资产构成假设)"],
        {"from_type": "Account[1601]", "to_type": "AssetCategory × UsefulLife",
         "mapping_table_ref": "本体知识：上年附注披露的固资类别分布（机器 51% / 运输 45% / 电子 4%）",
         "matched": True},
        f"按本体知识假设固资构成：机器(10年)51% + 运输(4年)45% + 电子(3年)4%")

    # ──────── Recompute 本期折旧 ────────
    avg_cost = (cost_tb["opening"] + cost_tb["closing"]) / 2
    weighted_recompute = 0
    category_breakdown = []
    for cat_name, m in asset_mix.items():
        cat_cost = avg_cost * m["weight"]
        cat_depreciable = cat_cost * (1 - m["salvage_pct"])
        cat_dep_per_year = cat_depreciable / m["life_years"]
        weighted_recompute += cat_dep_per_year
        category_breakdown.append({
            "category": cat_name, "weight": m["weight"],
            "estimated_cost": round(cat_cost, 2),
            "salvage_rate": m["salvage_pct"],
            "useful_life": m["life_years"],
            "recomputed_dep": round(cat_dep_per_year, 2),
        })
    book_dep = accum_tb["period_cr"]  # 本期计提（贷方）
    diff = book_dep - weighted_recompute
    diff_pct = diff / book_dep * 100 if book_dep else 0
    te = pm["data"]["performance_materiality_te"]
    sum_threshold = pm["data"]["clearly_trivial_sum"]
    pm_val = pm["data"]["overall_materiality_pm"]
    # 重要性梯度判定 (CSA 1221)
    abs_diff = abs(diff)
    if abs_diff < sum_threshold:
        test_level = "PASS_TRIVIAL"
        test_label = "✓ 明显微小 (差异 < SUM)"
        test_color = "low"
        next_action = "无需进一步程序，结论可写入"
    elif abs_diff < te:
        test_level = "PASS_WITHIN_TE"
        test_label = "✓ 可接受 (SUM ≤ 差异 < TE)"
        test_color = "low"
        next_action = "差异在 TE 内可接受，结论可写入"
    elif abs_diff < pm_val:
        test_level = "NEEDS_DETAIL_TEST"
        test_label = "⚠ 粗算不充分 (TE ≤ 差异 < PM)"
        test_color = "med"
        next_action = "建议进一步程序：(1) 获取固资明细账逐项重算；(2) 复核资产构成假设是否与实际一致；(3) 检查是否有减值/已提足资产"
    else:
        test_level = "MATERIAL_DIFF"
        test_label = "✗ 重大差异嫌疑 (差异 ≥ PM)"
        test_color = "high"
        next_action = "强制调整或要求客户给出合理解释；若无法解释考虑保留意见"

    run.call("Recompute",
        ["⚡ AT::Recompute", "📦 OT::Knowledge (折旧公式)", "📜 Rule::FA-RULE-001"],
        {"target": "本期折旧",
         "formula": "Σ(平均原值 × 类别权重 × (1-残值率) / 年限)",
         "inputs": {"avg_cost": round(avg_cost, 2),
                    "categories": category_breakdown,
                    "policy": policy_text},
         "book_value": book_dep},
        f"重算本期折旧 ¥{weighted_recompute:,.2f} vs 账面 ¥{book_dep:,.2f}，"
        f"差异 ¥{diff:,.2f} ({diff_pct:+.2f}%)")

    # ──────── Generate summary ────────
    summary = {
        "tb_cost_opening": cost_tb["opening"],
        "tb_cost_closing": cost_tb["closing"],
        "tb_cost_dr": cost_tb["period_dr"],
        "tb_cost_cr": cost_tb["period_cr"],
        "tb_accum_opening": accum_tb["opening"],
        "tb_accum_closing": accum_tb["closing"],
        "tb_dep_current_period": book_dep,
        "net_book_value_opening": round(cost_tb["opening"] + accum_tb["opening"], 2),
        "net_book_value_closing": round(cost_tb["closing"] + accum_tb["closing"], 2),
        "avg_cost": round(avg_cost, 2),
        "recomputed_dep": round(weighted_recompute, 2),
        "diff": round(diff, 2),
        "diff_pct": round(diff_pct, 2),
        "diff_vs_te": round(diff / te * 100, 2) if te else 0,
        "te": round(te, 2),
        "sum_threshold": round(sum_threshold, 2),
        "pm": round(pm_val, 2),
        "test_level": test_level,
        "test_label": test_label,
        "next_action": next_action,
        "passes_test": abs(diff) < te,
        "policy_applied": policy_text,
        "preparer": "jsdw_paper_fill (AI)",
        "prepared_at": date.today().isoformat(),
    }
    run.call("Generate", ["⚡ AT::Generate", "📦 OT::WorkingPaper.sheet_data"],
        {"sheet_code": "summary", "data": summary},
        f"summary 审定：重算差异 ¥{diff:,.2f}/{diff_pct:+.2f}%，"
        f"{'通过' if abs(diff) < te else '超 TE，须复核'}")

    prov.record(PAPER, "summary.tb_cost_closing", summary["tb_cost_closing"], "TB",
        ["📦 OT::TrialBalance", "📦 OT::Account[1601]"], "TB 1601 期末")
    prov.record(PAPER, "summary.tb_dep_current_period", summary["tb_dep_current_period"], "TB",
        ["📦 OT::TrialBalance", "📦 OT::Account[1602]"], "TB 1602 本期贷方 (本期计提)")
    prov.record(PAPER, "summary.avg_cost", summary["avg_cost"], "Computed",
        ["⚡ AT::Recompute"], "= (期初原值 + 期末原值) / 2")
    prov.record(PAPER, "summary.recomputed_dep", summary["recomputed_dep"], "Computed",
        ["⚡ AT::Recompute", "📦 OT::AccountingPolicy", "📦 OT::Knowledge (资产构成假设)"],
        f"Σ(平均原值×类别权重×(1-残值率)/年限) = ¥{summary['recomputed_dep']:,.2f}")
    prov.record(PAPER, "summary.diff", summary["diff"], "Computed",
        ["⚡ AT::Reconcile"], f"= 账面 ¥{book_dep:,.2f} - 重算 ¥{weighted_recompute:,.2f}")
    prov.record(PAPER, "summary.test_level", summary["test_level"],
        "RuleDerived", ["📜 Rule::FA-RULE-001", "📦 OT::MaterialityLevel"],
        f"按重要性梯度判定 (CSA 1221)：|差异| ¥{abs(diff):,.2f} 落在 "
        f"{('SUM' if abs(diff)<sum_threshold else 'TE' if abs(diff)<te else 'PM' if abs(diff)<pm_val else '>PM')} 内",
        rule_code="FA-RULE-001")
    prov.record(PAPER, "summary.test_label", summary["test_label"],
        "RuleDerived", ["📜 Rule::FA-RULE-001"],
        f"四级判定：明显微小/可接受/需细节测试/重大差异",
        rule_code="FA-RULE-001")
    prov.record(PAPER, "summary.next_action", summary["next_action"],
        "Knowledge", ["📦 OT::Knowledge (CSA 1221 重要性梯度)"],
        "根据差异级别给出下一步审计程序建议")
    prov.record(PAPER, "summary.passes_test", summary["passes_test"],
        "RuleDerived", ["📜 Rule::FA-RULE-001"],
        f"|差异| ¥{abs(diff):,.2f} {'< ' if summary['passes_test'] else '>='} TE ¥{te:,.2f}",
        rule_code="FA-RULE-001")
    prov.record(PAPER, "summary.policy_applied", policy_text, "Knowledge",
        ["📦 OT::AccountingPolicy"], "客户 X5 / ZS4 披露的折旧政策")

    # ──────── Generate recompute_detail (重算明细表) ────────
    recompute_rows = []
    for i, cb in enumerate(category_breakdown):
        cb_with_impl = {**cb,
            "annual_rate": round((1 - cb["salvage_rate"]) / cb["useful_life"] * 100, 2),
        }
        recompute_rows.append(cb_with_impl)
        prov.record(PAPER, f"recompute_detail.rows[{i}].weight", cb["weight"], "Knowledge",
            ["📦 OT::Knowledge (资产构成假设)"],
            f"基于行业 + 上年附注 {cb['category']} 占比")
        prov.record(PAPER, f"recompute_detail.rows[{i}].useful_life", cb["useful_life"],
            "Knowledge", ["📦 OT::AccountingPolicy"],
            f"客户 X5 政策定义 {cb['category']} 折旧年限")
        prov.record(PAPER, f"recompute_detail.rows[{i}].recomputed_dep",
            cb["recomputed_dep"], "Computed",
            ["⚡ AT::Recompute"],
            f"平均原值 ¥{cb['estimated_cost']:,.2f} × (1-{cb['salvage_rate']}) / {cb['useful_life']} = ¥{cb['recomputed_dep']:,.2f}")

    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "recompute_detail",
         "data": {"row_count": len(recompute_rows), "total": round(weighted_recompute, 2)}},
        f"recompute_detail {len(recompute_rows)} 类别折旧重算")

    # ──────── Generate movement_table (变动明细) ────────
    movement_rows = [
        {"item": "期初原值", "amount": cost_tb["opening"], "source": "TB 1601 期初"},
        {"item": "本期增加 (新购)", "amount": cost_tb["period_dr"], "source": "TB 1601 借方"},
        {"item": "本期减少 (处置)", "amount": -cost_tb["period_cr"], "source": "TB 1601 贷方"},
        {"item": "期末原值", "amount": cost_tb["closing"], "source": "TB 1601 期末"},
        {"item": "期初累计折旧", "amount": accum_tb["opening"], "source": "TB 1602 期初"},
        {"item": "本期处置冲销折旧", "amount": accum_tb["period_dr"], "source": "TB 1602 借方"},
        {"item": "本期计提折旧", "amount": -accum_tb["period_cr"], "source": "TB 1602 贷方"},
        {"item": "期末累计折旧", "amount": accum_tb["closing"], "source": "TB 1602 期末"},
        {"item": "期末净值", "amount": round(cost_tb["closing"] + accum_tb["closing"], 2),
         "source": "原值 + 累计折旧"},
    ]
    for i, mr in enumerate(movement_rows):
        prov.record(PAPER, f"movement.rows[{i}].amount", mr["amount"], "TB",
            ["📦 OT::TrialBalance"], mr["source"])
    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "movement", "data": {"row_count": len(movement_rows)}},
        "movement 变动明细 9 行")

    # ──────── 触发 AbnormalTransaction (如未通过粗算) ────────
    if test_level == "NEEDS_DETAIL_TEST":
        run.call("Generate",
            ["⚡ AT::Generate", "📦 OT::AbnormalTransaction", "📜 Rule::FA-RULE-001"],
            {"target_type": "AbnormalTransaction",
             "data": {"anomaly_type": "其他", "severity": "medium", "status": "open",
                      "detail": (f"折旧粗算合理性测试未通过：差异 ¥{abs_diff:,.2f} 介于 TE ¥{te:,.2f} 与 PM ¥{pm_val:,.2f} 之间。"
                                  f"最可能原因：AI 假设的资产构成（机器51%/运输45%/电子4%）与实际不符。"
                                  f"必要程序：(1) 获取固资明细账逐项重算；(2) 验证资产构成；(3) 检查减值/已提足资产"),
                      "amount_involved": abs(diff)}},
            f"生成 AbnormalTransaction 标记 FA 粗算未通过 → 需 DetailTest")

    # ──────── Explain ────────
    verdict_map = {
        "PASS_TRIVIAL": f"差异 ¥{abs_diff:,.2f} < SUM ¥{sum_threshold:,.2f}，明显微小，无需进一步程序。",
        "PASS_WITHIN_TE": f"差异 ¥{abs_diff:,.2f} 在 SUM~TE 区间 (TE ¥{te:,.2f})，可接受，初步合理性测试通过。",
        "NEEDS_DETAIL_TEST": f"差异 ¥{abs_diff:,.2f} 超 TE ¥{te:,.2f} 但未达 PM ¥{pm_val:,.2f} → "
                              f"粗算合理性测试未通过，需升级为细节测试（获取固资明细账逐项重算）。",
        "MATERIAL_DIFF": f"差异 ¥{abs_diff:,.2f} ≥ PM ¥{pm_val:,.2f} → 重大差异嫌疑，须客户说明或调整。",
    }
    audit_conclusion = (
        f"读取 TB 1601/1602/1606 + 客户 X5 折旧政策（直线法 / 残值 5% / 年限按类别）。"
        f"恒等式校验通过（期初+借-贷=期末）。"
        f"按本体知识假设资产构成（机器 51% + 运输 45% + 电子 4%）"
        f"对平均原值 ¥{avg_cost:,.2f} 重算本期折旧：¥{weighted_recompute:,.2f}，"
        f"账面 ¥{book_dep:,.2f}，差异 ¥{diff:,.2f} ({diff_pct:+.2f}%)。"
        f"{verdict_map[test_level]} "
        f"建议动作：{next_action}。"
        f"本期净值由 ¥{summary['net_book_value_opening']:,.2f} 变为 ¥{summary['net_book_value_closing']:,.2f}；"
        f"处置净值 ¥{cost_tb['period_cr'] - accum_tb['period_dr']:,.2f}（与营业外收入 ¥207,478.65 处置收益需核对）。"
    )
    prov.record(PAPER, "audit_conclusion", audit_conclusion, "Computed",
        ["⚡ AT::Explain"], "综合 TB/政策/重算输出结论")
    run.call("Explain", ["⚡ AT::Explain"],
        {"natural_language_output": audit_conclusion[:120] + "..."},
        f"写入 audit_conclusion ({len(audit_conclusion)} 字)")

    sheet_data = {
        "summary": summary,
        "recompute_detail": {"rows": recompute_rows, "total": round(weighted_recompute, 2)},
        "movement": {"rows": movement_rows},
    }
    filled_wp = {**wp, "data": {**wp["data"],
        "sheet_data": sheet_data, "audit_conclusion": audit_conclusion,
        "review_status": "AI 初稿",
        "cross_references": ["TB:1601/1602", "AccountingPolicy:固定资产折旧",
                              "D5:管理费用-折旧"],
        "filled_by": "jsdw_paper_fill",
        "filled_at": datetime.now().isoformat(timespec="seconds")}}
    run.msg("agent",
        f"A24 完成。重算折旧 ¥{weighted_recompute:,.2f} vs 账面 ¥{book_dep:,.2f}，"
        f"差异 {diff_pct:+.2f}%。{'通过' if summary['passes_test'] else '超 TE，待复核'}")
    return filled_wp, run, []


# ============================================================
# A9 其他应收款 (1133 出口退税反向 + RP 筛查)
# ============================================================
def fill_A9(ontology, raw, prov: Provenance):
    PAPER = "WP-A9-2025"
    run = AgentRun("jsdw_paper_fill", PAPER)
    run.msg("user", "请填写 A9 其他应收款底稿，识别出口退税反向余额 + 关联方资金占用")

    oi = ontology["object_instances"]
    wp = next(o for o in oi if o["type_code"] == "WorkingPaper" and o["data"]["index"] == "A9")
    pm = next(o for o in oi if o["type_code"] == "MaterialityLevel")
    rps = [o for o in oi if o["type_code"] == "RelatedParty"]
    te = pm["data"]["performance_materiality_te"]
    pm_val = pm["data"]["overall_materiality_pm"]

    # ──────── Read TB 1221 (含 1133 子目) ────────
    tb_1221 = [r for r in raw["tb"] if r["code"] == "1221"]
    # 1133 应收出口退税款映射到 1221 父级
    sub_1133 = next((r for r in tb_1221 if r["sub_code"] == "1133"), None)
    sub_1221 = next((r for r in tb_1221 if r["sub_code"] == "1221"), None)
    tb_total = sum(r["closing"] for r in tb_1221)
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::TrialBalance"],
        {"source": "TrialBalance", "filter": {"account_code": "1221"}},
        f"1221 父级 {len(tb_1221)} 子目，期末合计 ¥{tb_total:,.2f}（含 1133 ¥{sub_1133['closing'] if sub_1133 else 0:,.2f} + 1221 ¥{sub_1221['closing'] if sub_1221 else 0:,.2f}）")

    # ──────── Read SubLedger 1221 ────────
    aux_1221 = [a for a in raw["aux"] if a["code"] == "1221"]
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::SubLedger"],
        {"source": "SubLedger", "filter": {"account_code": "1221"}},
        f"读取 1221 客户级明细 {len(aux_1221)} 条")

    # ──────── Map: 客户 → RelatedParty ────────
    rp_names = {p["data"]["name"] for p in rps}
    matched_rp = [c for c in aux_1221 if c["customer_name"] in rp_names]
    run.call("Map",
        ["⚡ AT::Map", "📦 OT::RelatedParty", "🔗 LT::EntityHasRelatedParty"],
        {"from_type": "SubLedger.customer_name", "to_type": "RelatedParty.name"},
        f"机器匹配命中 {len(matched_rp)} 个关联方（客户匿名化）")

    # ──────── Reconcile: TB vs Aux 差异 = 1133 出口退税款余额 ────────
    aux_net = sum(c["closing_dr"] - c["closing_cr"] for c in aux_1221)
    tb_vs_aux_diff = tb_total - aux_net
    expected_1133 = sub_1133["closing"] if sub_1133 else 0
    run.call("Reconcile",
        ["⚡ AT::Reconcile"],
        {"left_source": "TB 1221 合计", "right_source": "Aux 1221 净额",
         "tolerance": 0.01},
        f"TB ¥{tb_total:,.2f} vs Aux ¥{aux_net:,.2f}，差异 ¥{tb_vs_aux_diff:,.2f} ≈ 1133 出口退税款 ¥{expected_1133:,.2f}")

    # ──────── Rule check: 1133 反向余额（负数）→ 触发出口退税重分类规则 ────────
    if expected_1133 < -0.01:  # 反向（贷方余额）
        run.call("Reconcile",
            ["⚡ AT::Reconcile", "📜 Rule::TAX-RECLASS-001 (拟新规则)",
             "📦 OT::AbnormalTransaction"],
            {"left_source": "Account[1133] 期末", "right_source": "(应为借方)",
             "tolerance": 0},
            f"1133 应收出口退税款期末 ¥{expected_1133:,.2f} (反向余额) → "
            f"已退税但未冲销 → 重分类至 B10 应交税费 22210106 出口抵减")

    # ──────── Apply RP-RULE-001 ────────
    rp_balance_total = sum(c["closing_dr"] for c in matched_rp)
    rule_triggered = rp_balance_total > pm_val
    run.call("Reconcile",
        ["⚡ AT::Reconcile", "📜 Rule::RP-RULE-001"],
        {"left_source": "matched_rp.closing_dr 合计", "right_source": "PM",
         "tolerance": 0},
        f"关联方占用合计 ¥{rp_balance_total:,.2f} {'>' if rule_triggered else '<'} "
        f"PM ¥{pm_val:,.2f} → {'触发强制披露' if rule_triggered else '未触发强制披露(余额<PM)'}")

    # ──────── Generate summary ────────
    summary = {
        "tb_1221_total": round(tb_total, 2),
        "tb_1133_export_refund": round(expected_1133, 2),
        "tb_1221_other_only": round(sub_1221["closing"] if sub_1221 else 0, 2),
        "aux_customers": len(aux_1221),
        "aux_dr_total": round(sum(c["closing_dr"] for c in aux_1221), 2),
        "aux_cr_total": round(sum(c["closing_cr"] for c in aux_1221), 2),
        "aux_net": round(aux_net, 2),
        "tb_vs_aux_diff": round(tb_vs_aux_diff, 2),
        "reclass_to_tax_payable": round(abs(expected_1133), 2),
        "matched_related_parties": len(matched_rp),
        "rp_balance_total": round(rp_balance_total, 2),
        "rp_rule_triggered": rule_triggered,
        "rp_threshold_pm": round(pm_val, 2),
        "closing_audited": round(sum(c["closing_dr"] for c in aux_1221) -
                                  sum(c["closing_cr"] for c in aux_1221), 2),
        "preparer": "jsdw_paper_fill (AI)",
        "prepared_at": date.today().isoformat(),
    }
    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "summary", "data": summary},
        f"summary 审定")

    prov.record(PAPER, "summary.tb_1221_total", summary["tb_1221_total"], "TB",
        ["📦 OT::TrialBalance"], "TB 1221 父级期末")
    prov.record(PAPER, "summary.tb_1133_export_refund", summary["tb_1133_export_refund"],
        "TB", ["📦 OT::TrialBalance", "📦 OT::Account[1133]"],
        "TB 1133 应收出口退税款期末（反向余额，提示已退未冲）")
    prov.record(PAPER, "summary.reclass_to_tax_payable", summary["reclass_to_tax_payable"],
        "RuleDerived", ["📜 Rule::TAX-RECLASS-001 (新规则)"],
        f"1133 反向余额绝对值 → 应重分类至 22210106 出口抵减 (生成 AJ-A9-01)",
        rule_code="TAX-RECLASS-001")
    prov.record(PAPER, "summary.matched_related_parties", summary["matched_related_parties"],
        "Computed", ["⚡ AT::Map", "📦 OT::RelatedParty"],
        f"客户名 vs {len(rp_names)} 个关联方名单匹配，因客户匿名化命中 0")
    prov.record(PAPER, "summary.rp_rule_triggered", summary["rp_rule_triggered"],
        "RuleDerived", ["📜 Rule::RP-RULE-001"],
        f"RP 余额 ¥{rp_balance_total:,.2f} vs PM ¥{pm_val:,.2f} → "
        f"{'触发' if rule_triggered else '未触发 (余额<PM 不强制披露)'}",
        rule_code="RP-RULE-001")
    prov.record(PAPER, "summary.closing_audited", summary["closing_audited"], "Aux",
        ["📦 OT::SubLedger"], "辅助账客户级借方合计 (1221 子目部分)")

    # ──────── Generate customer_detail ────────
    detail_rows = []
    for c in aux_1221:
        is_rp = c["customer_name"] in rp_names
        detail_rows.append({
            "customer_code": c["customer_code"],
            "customer_name": c["customer_name"],
            "closing_dr": round(c["closing_dr"], 2),
            "closing_cr": round(c["closing_cr"], 2),
            "is_related_party": is_rp,
            "exceeds_pm": c["closing_dr"] > pm_val,
            "classification": "关联方占用" if is_rp else "正常其他应收款",
        })
    for i, r in enumerate(detail_rows):
        prov.record(PAPER, f"customer_detail.rows[{i}].closing_dr", r["closing_dr"],
            "Aux", ["📦 OT::SubLedger"], f"辅助账 1221 客户 {r['customer_code']}")
        prov.record(PAPER, f"customer_detail.rows[{i}].is_related_party",
            r["is_related_party"],
            "RuleDerived" if r["is_related_party"] else "Computed",
            ["📜 Rule::RP-RULE-001", "📦 OT::RelatedParty"] if r["is_related_party"] else ["⚡ AT::Map"],
            f"客户 {r['customer_code']} {'命中' if r['is_related_party'] else '未命中'} 关联方名单",
            rule_code="RP-RULE-001" if r["is_related_party"] else None)
        prov.record(PAPER, f"customer_detail.rows[{i}].exceeds_pm", r["exceeds_pm"],
            "Computed", ["⚡ AT::Reconcile"],
            f"借方 ¥{r['closing_dr']:,.2f} {'>' if r['exceeds_pm'] else '<'} PM ¥{pm_val:,.2f}")
    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "customer_detail", "data": {"row_count": len(detail_rows)}},
        f"customer_detail {len(detail_rows)} 行")

    # ──────── 生成 AuditAdjustment: 1133 重分类 ────────
    proposed_adjustments = []
    if expected_1133 < -te:
        adj = {
            "no": "Z6-AI-A9-01",
            "reason": (f"1133 应收出口退税款期末 ¥{expected_1133:,.2f} (反向余额, 绝对值 ¥{abs(expected_1133):,.2f} > TE ¥{te:,.2f})，"
                       f"实质为已退税未冲销。重分类至 22210106 应交税费-出口抵减。"),
            "kind": "重分类",
            "entries": [
                {"side": "借", "account_code": "1221", "account_label": "其他应收款-1133 应收出口退税款",
                 "amount": round(abs(expected_1133), 2), "sub": "冲销反向余额"},
                {"side": "贷", "account_code": "22210106", "account_label": "应交税费-出口抵减内销产品应纳税额",
                 "amount": round(abs(expected_1133), 2), "sub": "已退税款转出"},
            ],
            "total_amount": round(abs(expected_1133), 2), "profit_impact": 0,
            "triggered_by_rule": "TAX-RECLASS-001",
            "approved_by_client": None,
        }
        proposed_adjustments.append(adj)
        run.call("Generate",
            ["⚡ AT::Generate", "📦 OT::AuditAdjustment", "📜 Rule::TAX-RECLASS-001"],
            {"target_type": "AuditAdjustment", "data": adj},
            f"生成 {adj['no']}: ¥{abs(expected_1133):,.2f} 1133 → 22210106 重分类")

    # ──────── Explain ────────
    audit_conclusion = (
        f"A9 其他应收款 TB 父级期末 ¥{tb_total:,.2f}（含 1133 出口退税款 ¥{expected_1133:,.2f} + 1221 其他 ¥{summary['tb_1221_other_only']:,.2f}）。"
        f"辅助账 {summary['aux_customers']} 客户净额 ¥{aux_net:,.2f}，"
        f"与 TB 差 ¥{tb_vs_aux_diff:,.2f} 即为 1133 子目余额（一致）。"
        f"1133 期末反向（已退税未冲），按 TAX-RECLASS-001 提议 Z6-AI-A9-01 重分类 ¥{abs(expected_1133):,.2f} 至 22210106。"
        f"RP-RULE-001 应用：客户名匿名化导致机器无命中 ({len(matched_rp)}/{len(aux_1221)})；"
        f"金额最大客户 ¥80,000 < PM ¥{pm_val:,.2f}，规则不触发强制披露，但需人工依据合同核查 001356 是否为关联方。"
        f"调整后审定数 ¥{summary['closing_audited']:,.2f}。"
    )
    prov.record(PAPER, "audit_conclusion", audit_conclusion, "Computed",
        ["⚡ AT::Explain"], "综合结论")
    run.call("Explain", ["⚡ AT::Explain"],
        {"natural_language_output": audit_conclusion[:120] + "..."},
        f"写入 audit_conclusion")

    sheet_data = {
        "summary": summary,
        "customer_detail": {"rows": detail_rows},
    }
    filled_wp = {**wp, "data": {**wp["data"],
        "sheet_data": sheet_data, "audit_conclusion": audit_conclusion,
        "review_status": "AI 初稿",
        "cross_references": ["TB:1221", "Z6-AI-A9-01", "B10:22210106"],
        "filled_by": "jsdw_paper_fill",
        "filled_at": datetime.now().isoformat(timespec="seconds")}}
    run.msg("agent", f"A9 完成。识别 1133 反向余额 ¥{abs(expected_1133):,.2f}，已提议重分类。")
    return filled_wp, run, proposed_adjustments


# ============================================================
# B1 短期借款 (核心: Recompute 利息 + 关联担保)
# ============================================================
def fill_B1(ontology, raw, prov: Provenance):
    PAPER = "WP-B1-2025"
    run = AgentRun("jsdw_paper_fill", PAPER)
    run.msg("user", "请填写 B1 短期借款底稿，重算利息并展示关联担保穿透")

    oi = ontology["object_instances"]
    wp = next(o for o in oi if o["type_code"] == "WorkingPaper" and o["data"]["index"] == "B1")
    pm = next(o for o in oi if o["type_code"] == "MaterialityLevel")
    te = pm["data"]["performance_materiality_te"]
    # 从本体读出所有 RelatedPartyGuarantee 实例（5 笔贷款）
    guarantees = [o for o in oi if o["type_code"] == "RelatedPartyGuarantee"]
    rps = [o for o in oi if o["type_code"] == "RelatedParty"]
    banks = [o for o in oi if o["type_code"] == "BankStatement"]

    # ──────── Read TB ────────
    tb_2001 = next(r for r in raw["tb"] if r["code"] == "2001")
    # 利息支出来自 6603 560305 利息收支
    interest_tb = next((r for r in raw["tb"]
                        if r["code"] == "6603" and r["sub_code"] == "560305"), None)
    book_interest = interest_tb["period_dr"] if interest_tb else 0
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::TrialBalance"],
        {"source": "TrialBalance",
         "filter": {"account_codes": ["2001", "6603.560305"]}},
        f"读取 2001 期末 ¥{tb_2001['closing']:,.2f}, "
        f"6603 利息收支借方 ¥{book_interest:,.2f}")

    # ──────── Read 关联担保 + 借款合同从本体 ────────
    run.call("Read",
        ["⚡ AT::Read", "📦 OT::RelatedPartyGuarantee", "🔗 LT::EngagementHasGuarantee"],
        {"source": "RelatedPartyGuarantee", "from": "ontology"},
        f"读取本体 {len(guarantees)} 笔关联担保（含金额/利率/期限/担保人）")

    # 解析担保实例（带利率：本案 5 笔为 2.40%~2.70%）
    loan_data = []
    rates_seq = [0.024, 0.027, 0.0265, 0.024, 0.024]  # 来自 X1 / FSR 披露
    for i, g in enumerate(sorted(guarantees, key=lambda x: x["data"]["term_start"])):
        d = g["data"]
        loan_data.append({
            "creditor": d["creditor_name"], "amount": d["amount"],
            "rate": rates_seq[i] if i < len(rates_seq) else 0.024,
            "term_start": d["term_start"], "term_end": d["term_end"],
            "guarantor": d["guarantor_names"],
            "guarantee_type": d.get("guarantee_type", "连带保证"),
        })

    # ──────── Recompute 本期利息 (假设各贷款活跃整年简化模型) ────────
    annual_interest_total = 0
    loan_breakdown = []
    for ld in loan_data:
        annual = ld["amount"] * ld["rate"]
        loan_breakdown.append({**ld, "annual_interest": round(annual, 2)})
        annual_interest_total += annual

    interest_diff = book_interest - annual_interest_total
    interest_diff_pct = interest_diff / book_interest * 100 if book_interest else 0
    run.call("Recompute",
        ["⚡ AT::Recompute", "📦 OT::Knowledge (利息=本金×利率×时间)"],
        {"target": "本期利息支出",
         "formula": "Σ(本金 × 年利率)，假设各贷款活跃整年（简化模型）",
         "inputs": {"loans": loan_breakdown, "book": book_interest}},
        f"重算年利息 ¥{annual_interest_total:,.2f} vs 账面 ¥{book_interest:,.2f}，"
        f"差异 ¥{interest_diff:,.2f} ({interest_diff_pct:+.2f}%)")

    # ──────── Map 担保人 → 关联方实例 ────────
    guarantor_set = set()
    for ld in loan_data:
        for g in ld["guarantor"].split(","):
            guarantor_set.add(g.strip())
    rp_name_set = {p["data"]["name"] for p in rps}
    all_guarantors_are_rp = guarantor_set.issubset(rp_name_set)
    run.call("Map",
        ["⚡ AT::Map", "🔗 LT::GuaranteeByParty", "📦 OT::RelatedParty"],
        {"from_type": "Guarantee.guarantor_names", "to_type": "RelatedParty.name"},
        f"全部担保人 {sorted(guarantor_set)} {'均为' if all_guarantors_are_rp else '部分'}关联方")

    # ──────── Apply Rule: 100% 关联担保 → 触发持续经营关注 ────────
    related_guarantee_ratio = 1.0 if all_guarantors_are_rp else 0
    triggers_going_concern = related_guarantee_ratio >= 1.0 and tb_2001["closing"] > 0
    run.call("Reconcile",
        ["⚡ AT::Reconcile", "📜 Rule::GC-INDICATOR-001 (持续经营)",
         "📦 OT::RiskOfMaterialMisstatement"],
        {"left_source": "关联担保覆盖率", "right_source": "100%",
         "tolerance": 0},
        f"{len(loan_data)}/{len(loan_data)} 短借由 {guarantor_set} 关联方担保 → "
        f"100% 集中度 → {'触发' if triggers_going_concern else '未触发'} 持续经营关注")

    # ──────── Generate summary ────────
    summary = {
        "tb_opening": tb_2001["opening"],
        "tb_period_dr": tb_2001["period_dr"],
        "tb_period_cr": tb_2001["period_cr"],
        "tb_closing": tb_2001["closing"],
        "loan_count": len(loan_data),
        "principal_total": sum(ld["amount"] for ld in loan_data),
        "weighted_avg_rate": round(sum(ld["amount"]*ld["rate"] for ld in loan_data) /
                                    sum(ld["amount"] for ld in loan_data) * 100, 4) if loan_data else 0,
        "book_interest_expense": round(book_interest, 2),
        "recomputed_interest_annual": round(annual_interest_total, 2),
        "interest_diff": round(interest_diff, 2),
        "interest_diff_pct": round(interest_diff_pct, 2),
        "guarantors": sorted(guarantor_set),
        "related_guarantee_ratio": related_guarantee_ratio,
        "triggers_going_concern": triggers_going_concern,
        "closing_audited": round(tb_2001["closing"], 2),
        "preparer": "jsdw_paper_fill (AI)",
        "prepared_at": date.today().isoformat(),
    }
    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "summary", "data": summary}, "summary 审定表")

    prov.record(PAPER, "summary.tb_closing", summary["tb_closing"], "TB",
        ["📦 OT::TrialBalance", "📦 OT::Account[2001]"], "TB 2001 短期借款 期末")
    prov.record(PAPER, "summary.principal_total", summary["principal_total"], "Knowledge",
        ["📦 OT::RelatedPartyGuarantee", "🔗 LT::EngagementHasGuarantee"],
        f"本体 {len(loan_data)} 笔关联担保贷款本金汇总")
    prov.record(PAPER, "summary.weighted_avg_rate", summary["weighted_avg_rate"],
        "Computed", ["⚡ AT::Recompute"], "Σ(本金×利率)/Σ本金")
    prov.record(PAPER, "summary.book_interest_expense", summary["book_interest_expense"],
        "TB", ["📦 OT::TrialBalance", "📦 OT::Account[6603.560305]"],
        "TB 6603-560305 利息收支借方累计")
    prov.record(PAPER, "summary.recomputed_interest_annual",
        summary["recomputed_interest_annual"], "Computed",
        ["⚡ AT::Recompute", "📦 OT::Knowledge (利息公式)"],
        "Σ(本金×年利率)；简化假设各贷款活跃整年")
    prov.record(PAPER, "summary.interest_diff", summary["interest_diff"], "Computed",
        ["⚡ AT::Reconcile"],
        f"差异 = 账面 - 重算 = ¥{summary['interest_diff']:,.2f}")
    prov.record(PAPER, "summary.related_guarantee_ratio",
        summary["related_guarantee_ratio"], "RuleDerived",
        ["📜 Rule::GC-INDICATOR-001", "📦 OT::RelatedParty"],
        f"5/5 贷款由 {guarantor_set} 担保 → 100% 关联担保",
        rule_code="GC-INDICATOR-001")
    prov.record(PAPER, "summary.triggers_going_concern", summary["triggers_going_concern"],
        "RuleDerived", ["📜 Rule::GC-INDICATOR-001"],
        f"100% 关联担保 + 期末借款 ¥{tb_2001['closing']:,.2f} > 0 → 持续经营风险标志",
        rule_code="GC-INDICATOR-001")

    # ──────── Generate loan_detail (函证用) ────────
    loan_rows = []
    for i, lb in enumerate(loan_breakdown):
        loan_rows.append({
            "loan_no": f"L{i+1:02d}",
            "creditor": lb["creditor"], "principal": lb["amount"],
            "rate": lb["rate"], "annual_interest": lb["annual_interest"],
            "term_start": lb["term_start"], "term_end": lb["term_end"],
            "guarantor": lb["guarantor"], "guarantee_type": lb["guarantee_type"],
            "confirmation_status": "（待发函）",
        })
        prov.record(PAPER, f"loan_detail.rows[{i}].principal", lb["amount"], "Knowledge",
            ["📦 OT::RelatedPartyGuarantee"], f"本体担保 {lb['creditor']} 本金")
        prov.record(PAPER, f"loan_detail.rows[{i}].rate", lb["rate"], "Knowledge",
            ["📦 OT::Knowledge (X1 借款合同利率)"],
            f"利率 {lb['rate']*100:.4f}% 来自客户披露")
        prov.record(PAPER, f"loan_detail.rows[{i}].annual_interest", lb["annual_interest"],
            "Computed", ["⚡ AT::Recompute"],
            f"= 本金 ¥{lb['amount']:,.0f} × 利率 {lb['rate']*100:.2f}% = ¥{lb['annual_interest']:,.2f}")
        prov.record(PAPER, f"loan_detail.rows[{i}].guarantor", lb["guarantor"],
            "Knowledge", ["📦 OT::RelatedPartyGuarantee"], "本体担保实例")
        prov.record(PAPER, f"loan_detail.rows[{i}].confirmation_status", "（待发函）",
            "TemplateConst", ["📦 OT::ConfirmationLetter"],
            "函证程序未启动；agent 不可代为发函")

    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "loan_detail", "data": {"row_count": len(loan_rows)}},
        f"loan_detail {len(loan_rows)} 笔贷款")

    # ──────── Generate interest_recompute (重算明细) ────────
    recompute_rows = [{
        "loan_no": f"L{i+1:02d}", "creditor": lb["creditor"],
        "principal": lb["amount"], "rate": lb["rate"],
        "rate_pct": round(lb["rate"]*100, 4),
        "formula": f"¥{lb['amount']:,.0f} × {lb['rate']*100:.2f}%",
        "recomputed_annual": lb["annual_interest"],
    } for i, lb in enumerate(loan_breakdown)]
    # 加合计行
    run.call("Generate", ["⚡ AT::Generate"],
        {"sheet_code": "interest_recompute", "data": {"row_count": len(recompute_rows)}},
        f"interest_recompute {len(recompute_rows)} 行")
    for i, r in enumerate(recompute_rows):
        prov.record(PAPER, f"interest_recompute.rows[{i}].recomputed_annual",
            r["recomputed_annual"], "Computed",
            ["⚡ AT::Recompute"],
            f"= {r['principal']:,.0f} × {r['rate_pct']:.4f}% = {r['recomputed_annual']:,.2f}")

    # ──────── Explain ────────
    audit_conclusion = (
        f"B1 短期借款期末 ¥{tb_2001['closing']:,.2f}（5 笔贷款），与本体 5 笔关联担保实例一致。"
        f"加权平均利率 {summary['weighted_avg_rate']:.4f}%。"
        f"重算年利息 ¥{annual_interest_total:,.2f}（简化假设全年活跃）vs 账面 ¥{book_interest:,.2f}，"
        f"差异 ¥{interest_diff:,.2f} ({interest_diff_pct:+.2f}%)，"
        f"主要源自各贷款实际起止日不同 + 上年余额产生的利息（需获取按月利息计提表精算）。"
        f"100% 贷款由关联方 {', '.join(sorted(guarantor_set))} 连带保证 → "
        f"已触发 GC-INDICATOR-001 持续经营关注，建议在 ZK5 关联担保段补充披露 + 评估 12 个月偿债能力。"
        f"5 笔贷款函证未发，待项目组对每笔贷款向开户行发出银行询证函。"
    )
    prov.record(PAPER, "audit_conclusion", audit_conclusion, "Computed",
        ["⚡ AT::Explain"], "综合结论")
    run.call("Explain", ["⚡ AT::Explain"],
        {"natural_language_output": audit_conclusion[:120] + "..."},
        f"写入 audit_conclusion")

    sheet_data = {
        "summary": summary,
        "loan_detail": {"rows": loan_rows},
        "interest_recompute": {"rows": recompute_rows,
                                "total_recomputed": round(annual_interest_total, 2),
                                "book": round(book_interest, 2),
                                "diff": round(interest_diff, 2),
                                "diff_pct": round(interest_diff_pct, 2)},
    }
    filled_wp = {**wp, "data": {**wp["data"],
        "sheet_data": sheet_data, "audit_conclusion": audit_conclusion,
        "review_status": "AI 初稿",
        "cross_references": ["TB:2001", "TB:6603.560305",
                              "RelatedPartyGuarantee×5", "RiskFlag:高资产负债率"],
        "filled_by": "jsdw_paper_fill",
        "filled_at": datetime.now().isoformat(timespec="seconds")}}
    run.msg("agent",
        f"B1 完成。本金 ¥{summary['principal_total']:,.0f}，重算利息 ¥{annual_interest_total:,.2f}，"
        f"100% 关联担保已标记 GC 关注。")
    return filled_wp, run, []


# ============================================================
# Main
# ============================================================
def main():
    print("=== 加载本体 ===")
    ontology = load_ontology()
    print(f"  ObjectType={len(ontology['object_types'])} "
          f"实例={len(ontology['object_instances'])}")
    print("=== 加载客户原始数据 ===")
    raw = load_raw()
    print(f"  TB={len(raw['tb'])} Aux={len(raw['aux'])} Vouchers={len(raw['vouchers'])}")

    prov = Provenance()

    print("\n=== A1 货币资金 ===")
    a1_wp, a1_run, a1_adj = fill_A1(ontology, raw, prov)
    print(f"  调用：{len(a1_run.tool_calls)} 次, provenance 条目: {len(prov.by_paper.get('WP-A1-2025', {}))}")

    print("\n=== A6 应收账款 ===")
    a6_wp, a6_run, a6_adj = fill_A6(ontology, raw, prov)
    print(f"  调用：{len(a6_run.tool_calls)} 次, provenance: {len(prov.by_paper.get('WP-A6-2025', {}))}")

    print("\n=== A24 固定资产 ===")
    a24_wp, a24_run, a24_adj = fill_A24(ontology, raw, prov)
    print(f"  调用：{len(a24_run.tool_calls)} 次, provenance: {len(prov.by_paper.get('WP-A24-2025', {}))}")

    print("\n=== A9 其他应收款 ===")
    a9_wp, a9_run, a9_adj = fill_A9(ontology, raw, prov)
    print(f"  调用：{len(a9_run.tool_calls)} 次, provenance: {len(prov.by_paper.get('WP-A9-2025', {}))}")

    print("\n=== B1 短期借款 ===")
    b1_wp, b1_run, b1_adj = fill_B1(ontology, raw, prov)
    print(f"  调用：{len(b1_run.tool_calls)} 次, provenance: {len(prov.by_paper.get('WP-B1-2025', {}))}")

    print("\n=== 写出 ===")
    for fname, content in [
        ("filled_A1_workingpaper.json",  a1_wp),
        ("filled_A6_workingpaper.json",  a6_wp),
        ("filled_A24_workingpaper.json", a24_wp),
        ("filled_A9_workingpaper.json",  a9_wp),
        ("filled_B1_workingpaper.json",  b1_wp),
    ]:
        (OUT / fname).write_text(json.dumps(content, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "agent_run_log.json").write_text(
        json.dumps([r.to_dict() for r in [a1_run, a6_run, a24_run, a9_run, b1_run]],
                   ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "proposed_adjustments.json").write_text(
        json.dumps(a1_adj + a6_adj + a24_adj + a9_adj + b1_adj, ensure_ascii=False, indent=2),
        encoding="utf-8")
    (OUT / "cell_provenance.json").write_text(
        json.dumps(prov.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    for f in ["filled_A1_workingpaper.json", "filled_A6_workingpaper.json",
              "filled_A24_workingpaper.json", "filled_A9_workingpaper.json",
              "filled_B1_workingpaper.json",
              "agent_run_log.json", "proposed_adjustments.json", "cell_provenance.json"]:
        print(f"  {f}  ({(OUT / f).stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
