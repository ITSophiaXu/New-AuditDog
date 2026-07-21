from io import BytesIO
import unittest

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine, select

from app.annual_audit.router import (
    _fill_project_papers,
    _paper_workbook_bytes,
    _task_templates_for_project,
    _validate_materiality,
    build_sheet_data,
)
from app.models import ObjectInstance


class AnnualAuditTests(unittest.TestCase):
    def test_materiality_requires_descending_thresholds(self) -> None:
        _validate_materiality({"pm": 100_000, "te": 75_000, "trivial_threshold": 5_000})
        with self.assertRaises(HTTPException):
            _validate_materiality({"pm": 75_000, "te": 100_000, "trivial_threshold": 5_000})

    def test_subject_workpaper_contains_core_sheets(self) -> None:
        sheet_data = build_sheet_data(
            {
                "index": "A6",
                "name": "Accounts receivable",
                "stage": "execution",
                "cycle": "revenue",
                "kind": "subject",
            },
            {"pm": 100_000, "te": 75_000, "preparer": "Annual Audit Agent"},
        )

        self.assertIn("summary", sheet_data)
        self.assertIn("ledger_detail", sheet_data)
        self.assertIn("audit_procedures", sheet_data)
        self.assertIn("aging_analysis", sheet_data)

    def test_non_demo_project_uses_generic_review_tasks(self) -> None:
        test_engine = create_engine("sqlite://")
        self.addCleanup(test_engine.dispose)
        SQLModel.metadata.create_all(test_engine)
        with Session(test_engine) as session:
            engagement = ObjectInstance(
                type_code="Engagement",
                display_name="Test engagement",
                data={
                    "code": "ENG-TEST-2025",
                    "use_demo_data": False,
                    "pm": 100_000,
                    "te": 75_000,
                    "trivial_threshold": 5_000,
                    "materiality_basis": "Revenue",
                },
            )
            session.add(engagement)
            session.add(ObjectInstance(
                type_code="UploadedMaterial",
                display_name="trial_balance.csv",
                data={
                    "engagement_code": "ENG-TEST-2025",
                    "category": "account_set",
                    "filename": "trial_balance.csv",
                },
            ))
            session.commit()

            templates = _task_templates_for_project(session, engagement)

        self.assertEqual(
            {template["task_key"] for template in templates},
            {
                "Y3-MATERIALITY",
                "TB1-DATA-COMPLETENESS",
                "X7-PBC-EVIDENCE",
                "Z5-OPINION",
            },
        )

    def test_safe_rerun_preserves_existing_workpaper_content(self) -> None:
        test_engine = create_engine("sqlite://")
        self.addCleanup(test_engine.dispose)
        SQLModel.metadata.create_all(test_engine)
        with Session(test_engine) as session:
            engagement = ObjectInstance(
                type_code="Engagement",
                display_name="Test engagement",
                data={
                    "code": "ENG-SAFE-RERUN",
                    "client_name": "Test client",
                    "accounting_standard": "Enterprise Accounting Standards",
                    "pm": 100_000,
                    "te": 75_000,
                    "trivial_threshold": 5_000,
                    "use_demo_data": False,
                },
            )
            session.add(engagement)
            session.commit()
            session.refresh(engagement)

            _fill_project_papers(session, engagement, preserve_existing=True)
            paper = next(
                item
                for item in session.exec(select(ObjectInstance))
                if item.type_code == "WorkingPaper"
                and (item.data or {}).get("engagement_code") == "ENG-SAFE-RERUN"
                and (item.data or {}).get("index") == "A6"
            )
            data = dict(paper.data or {})
            sheet_data = dict(data["sheet_data"])
            summary = dict(sheet_data["summary"])
            summary["manual_marker"] = "keep"
            sheet_data["summary"] = summary
            data["sheet_data"] = sheet_data
            paper.data = data
            session.add(paper)
            session.commit()

            filled = _fill_project_papers(session, engagement, preserve_existing=True)
            session.refresh(paper)

        self.assertEqual(filled, 0)
        self.assertEqual((paper.data or {})["template_code"], "TPL-AA-FY2025")
        self.assertEqual((paper.data or {})["sheet_data"]["summary"]["manual_marker"], "keep")

    def test_workpaper_exports_as_xlsx(self) -> None:
        engagement = ObjectInstance(
            type_code="Engagement",
            display_name="Test engagement",
            data={"period_end": "2025-12-31"},
        )
        paper = ObjectInstance(
            type_code="WorkingPaper",
            display_name="A6 Accounts receivable",
            data={
                "sheet_data": {
                    "summary": {"closing_balance": 100},
                    "ledger_detail": [{"customer": "A", "balance": 100}],
                },
            },
        )

        content = _paper_workbook_bytes(paper, engagement)
        workbook = load_workbook(BytesIO(content), read_only=True)

        self.assertEqual(workbook.sheetnames, ["summary", "ledger_detail"])


if __name__ == "__main__":
    unittest.main()
